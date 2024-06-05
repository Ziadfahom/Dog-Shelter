import json
import math
from collections import defaultdict, Counter
from datetime import datetime, time
from functools import wraps

from django.contrib.auth import REDIRECT_FIELD_NAME
from django.utils.timezone import localtime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.db.models import Prefetch
from django.db.models import Value, CharField, Count, Q
from django.db.models.functions import Concat
from django.shortcuts import render, redirect, get_object_or_404, resolve_url
from django.contrib.auth import authenticate, login, logout, get_user_model, update_session_auth_hash
from django.contrib.auth.models import Group
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.http import HttpResponseBadRequest, HttpResponseRedirect, JsonResponse, HttpResponse, \
    HttpResponseNotAllowed, HttpResponseForbidden
from .filters import DogFilter
from .forms import SignUpForm, AddDogForm, UpdateUserForm, ProfileUpdateForm, TreatmentForm, EntranceExaminationForm, \
    DogPlacementForm, ObservesForm, ObservationForm, DogStanceForm, LoginForm, NewsForm, OwnerForm
from .models import *
from django.conf import settings
import os
from django.core import serializers
from django.core.paginator import Paginator, EmptyPage
from django.template.loader import render_to_string
from datetime import date, timedelta
from django.db import IntegrityError, transaction
from io import BytesIO
from .serializers import DogSerializer
from django.views.decorators.http import require_POST
from django.core.serializers.json import DjangoJSONEncoder
from django.urls import reverse
from django.forms import formset_factory
from .forms import PollForm, ChoiceForm, ChoiceFormSet

# Location of the default User profile picture if they don't have a picture
DEFAULT_IMAGE_SOURCE = '/profile_pictures/default.jpg'

# Location of the default Dog picture if they don't have a picture
DEFAULT_DOG_IMAGE_SOURCE = '/dog_pictures/default_dog.jpg'

# CONSTANTS
# Number of Dogs/Users/Table Entries we want displayed in a single page in the Paginator
ENTRIES_PER_PAGE = 10


# View for handling switching between Israel and Italy branches
def toggle_branch(request):
    current_branch = request.session.get('branch', 'Israel')
    new_branch = 'Italy' if current_branch == 'Israel' else 'Israel'
    request.session['branch'] = new_branch
    return redirect('dogs_app:home')


# View for setting the branch to 'Italy'
def set_italy_branch(request):
    request.session['branch'] = 'Italy'
    return redirect('dogs_app:home')


# View for setting the branch to 'Israel'
def set_israel_branch(request):
    request.session['branch'] = 'Israel'
    return redirect('dogs_app:home')


# Helper function to get the user's current branch object (Israel/Italy)
def get_current_branch(request):
    # Get the current Branch (Israel/Italy)
    branch_name = request.session.get('branch', 'Israel')  # Default to Israel
    branch = Branch.objects.get(branchName=branch_name if branch_name else 'Israel')  # Get the branch object
    return branch


# Helper function to check if the user is logged in and is an admin
def superuser_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "You must be logged in as an admin to view this page.")
            return redirect(f"{resolve_url(settings.LOGIN_URL)}?{REDIRECT_FIELD_NAME}={request.path}")
        if not request.user.is_superuser:
            messages.error(request, "You must be an admin to view this page.")
            return redirect('dogs_app:home')
        return view_func(request, *args, **kwargs)

    return _wrapped_view


# Helper function to check if the user is logged in and is a veterinarian
def vet_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "You must be logged in as a veterinarian to view this page.")
            return redirect(f"{resolve_url(settings.LOGIN_URL)}?{REDIRECT_FIELD_NAME}={request.path}")
        if not request.user.groups.filter(name="Vet").exists() and not request.user.is_superuser:
            messages.error(request, "You must be a veterinarian to view this page.")
            return redirect('dogs_app:home')
        return view_func(request, *args, **kwargs)

    return _wrapped_view


# Return True if the user is validated as a Vet user
def valid_vet_user(request):
    if request.user.is_authenticated and (request.user.groups.filter(name='Vet').exists() or request.user.is_superuser):
        return True
    else:
        return False


# Return True if the user is validated as a registered Regular ('Viewer') user
def valid_registered_regular_user(request):
    if request.user.is_authenticated and (request.user.groups.filter(name='Vet').exists() or request.user.groups.filter(name='Viewer').exists() or request.user.is_superuser):
        return True
    else:
        return False


# Helper function to check if the user is logged in and is a regular (registered) user
def regular_user_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "You must be logged in to view this page.")
            return redirect(f"{resolve_url(settings.LOGIN_URL)}?{REDIRECT_FIELD_NAME}={request.path}")
        if not request.user.groups.filter(name="Viewer").exists() and not request.user.groups.filter(name="Vet").exists() and not request.user.is_superuser:
            messages.error(request, "You must be logged in to view this page.")
            return redirect('dogs_app:home')
        return view_func(request, *args, **kwargs)

    return _wrapped_view


# Display the homepage with the appropriate branch's data
def home_view(request):
    # Get the current Branch
    branch = get_current_branch(request)

    # Get the total number of dogs
    total_dogs = Dog.objects.filter(branch=branch).count()

    # Get the number of dogs that have received toy treatments for the selected branch
    toy_treatment_dogs = Dog.objects.filter(branch=branch).filter(observers__observation__isKong='Y').distinct().count()

    # Get all the website News to display them in descending order
    news_items = News.objects.filter(branch=branch).order_by('-created_at')[:3]

    # Get all polls to display them in descending order
    poll_items = Poll.objects.filter(branch=branch).order_by('-pub_date')[:3]  # Ordering by pub_date

    if request.method == 'POST':
        # Handle user login
        username = request.POST['username']
        password = request.POST['password']
        users = get_user_model()

        if users.objects.filter(username=username).exists():
            try:
                user = authenticate(request, username=username, password=password)
                if user is not None:
                    login(request, user=user)
                    messages.success(request, message='You have successfully logged in!')
                    return redirect('dogs_app:home')
                else:
                    messages.error(request, message='The password is incorrect. Please try again...')
                    return redirect('dogs_app:home')
            except Exception as e:
                messages.error(request, message=f'An error occurred during login: {e}')
                return redirect('dogs_app:home')
        else:
            messages.error(request, message='The username does not exist. Please try again...')
            return redirect('dogs_app:home')
    else:
        role = get_user_role(request.user) if request.user.is_authenticated else ""
        context = {
            'total_dogs': total_dogs,
            'toy_treatment_dogs': toy_treatment_dogs,
            'news_items': news_items,
            'poll_items': poll_items,
            'role': role
        }
        return render(request, 'home.html', context=context)


# Logout Users view for displaying a user-logout option if they're already logged in
def logout_user_view(request):
    # Get the current Branch, make sure it remains the same after logout
    branch = get_current_branch(request)
    logout(request)
    messages.success(request, message='You have been logged out..')
    set_israel_branch(request) if branch.branchName == 'Israel' else set_italy_branch(request)
    return redirect('dogs_app:home')


def login_user_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        users = get_user_model()

        # Check if username exists
        if users.objects.filter(username=username).exists():

            try:
                # Authenticate
                user = authenticate(request, username=username, password=password)

                # If correct username+password combination
                if user is not None and user.is_authenticated:
                    login(request, user=user)
                    messages.success(request, message='You have successfully logged in!')
                    next_url = request.POST.get('next') or 'dogs_app:home'
                    return redirect(next_url)
                else:
                    # Username exists but password was incorrect
                    messages.error(request, message='The password is incorrect. Please try again...')
                    return redirect('dogs_app:home')
            except Exception as e:
                messages.error(request, message=f'An error occurred during login: {e}')
                return redirect('dogs_app:home')
        else:
            # Username does not exist
            messages.error(request, message='The username does not exist. Please try again...')
            return redirect('dogs_app:home')
    else:
        form = LoginForm()
    return render(request, 'account/login.html', {'form': form})


# User Registration view for new users
def register_user_view(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        profile_form = ProfileUpdateForm(request.POST, request.FILES)
        if form.is_valid() and profile_form.is_valid():
            user = form.save()
            # Add User to the Viewers Group
            viewers_group = Group.objects.get(name="Viewer")
            user.groups.add(viewers_group)

            # Check if a profile already exists for the user
            profile = Profile.objects.filter(user=user).first()
            if not profile:
                profile = Profile.objects.create(user=user)

            profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)
            profile_form.save()

            username = form.cleaned_data['username']
            password = form.cleaned_data['password1']
            user = authenticate(request, username=username, password=password)
            login(request, user)

            messages.success(request, f"Welcome, {username} and thank you for signing up! "
                                      "You're now a member of our Dogswatch community.")
            return redirect('dogs_app:home')
    else:
        form = SignUpForm()
        profile_form = ProfileUpdateForm()

    return render(request, 'account/register.html', {'form': form, 'profile_form': profile_form})


@login_required
# User password reset view
def change_password(request):
    # User trying to submit the form
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        # The details are valid
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important! Re-logins the user after password change
            messages.success(request, 'Your password was successfully updated!')
            return redirect('dogs_app:home')
        # Issues with the filled details
        else:
            messages.error(request, 'Please correct the error below')
    # User is trying to open the change password page
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'account/change_password.html', {
        'form': form,
    })


@superuser_required
# View for adding website news to the homepage only for admins
def add_news(request):
    if request.method == 'POST':
        # Get the current Branch
        branch = get_current_branch(request)

        news_title = request.POST.get('title')
        news_content = request.POST.get('content')
        News.objects.create(title=news_title, content=news_content, branch=branch)
        messages.success(request, 'News has been added successfully!')
        return redirect('dogs_app:home')
    else:
        form = NewsForm()
        context = {
            'form': form
        }
    return render(request, 'add_news.html', context)


@regular_user_required
# Dog record page, displays the details for a single dog
def dog_record_view(request, pk):
    # Check if the user is logged in
    if request.user.is_authenticated:
        # Look up and save the dog's record and all relevant data of that dog
        dog_record = Dog.objects.select_related('owner').prefetch_related(
            'treatment_set',
            'entranceexamination_set',
            'observers__observation_set',
            'dogplacement_set__kennel',  # For Kennels related to DogPlacement's kennel
            'observers__observation_set__dogstance_set',  # For DogStances related to Observations
        ).get(dogID=pk)

        # Get the current Branch from the Dog entity to determine the appropriate chart data to fetch
        branch = dog_record.branch.branchName

        # Validate that the user is in the correct branch
        session_branch = get_current_branch(request).branchName
        if session_branch != branch:
            messages.error(request, "You must be in the correct branch to view this dog...")
            return redirect('dogs_app:home')

        # Handle displaying the summarization for the dog below the picture
        # Find the oldest observation date
        oldest_observation = Observation.objects.filter(
            observes__dog=dog_record
        ).order_by('obsDateTime').first()

        # Adjust the start date to the earliest observation if available
        if oldest_observation:
            first_date = localtime(oldest_observation.obsDateTime).date()
        else:
            first_date = date.today()

        # Get the first day of that month for the weekly map
        first_weekly_date = first_date.replace(day=1)

        last_date = date.today()

        # Prepare the heatmap data for each year, by days
        daily_heatmap_data = {}
        # Initialize dictionary for isKong counts
        kong_daily_counts = defaultdict(int)
        if branch == 'Italy':  # Italy branch has additional fields to prepare: isDog, isHuman, and no Stances
            is_dog_daily_counts = defaultdict(int)
            is_human_daily_counts = defaultdict(int)
            no_dog_stance_counts = defaultdict(int)

        # Initialize dictionary for weekly heatmap data depending on the branch
        if branch == 'Italy':   # [total count, kong count, dog count, human count, no stance count]
            weekly_heatmap_data = defaultdict(lambda: defaultdict(lambda: [0, 0, 0, 0, 0]))
        else:                   # [total count, kong count]
            weekly_heatmap_data = defaultdict(lambda: defaultdict(lambda: [0, 0]))

        # Fetch all observations for the dog in one query
        local_tz = pytz.timezone('Asia/Jerusalem')  # Fetch all observations for the dog in UTC
        utc_first_date = timezone.make_aware(datetime.combine(first_date, time.min), local_tz).astimezone(pytz.utc)
        utc_last_date = timezone.make_aware(datetime.combine(last_date, time.max), local_tz).astimezone(pytz.utc)

        # Fetch observations in UTC
        observations = Observation.objects.filter(
            observes__dog=dog_record,
            obsDateTime__range=(utc_first_date, utc_last_date)
        )

        # Convert obsDateTime to local timezone and count occurrences per day followed by per week
        observation_counts = defaultdict(int)
        for obs in observations:
            local_date = timezone.localtime(obs.obsDateTime, local_tz).date()
            observation_counts[local_date] += 1
            if obs.isKong == 'Y':
                kong_daily_counts[local_date] += 1
            if branch == 'Italy':  # Italy branch has additional fields to prepare
                if obs.isDog == 'Y':
                    is_dog_daily_counts[local_date] += 1  # Ignore interpreter warning, the if statement is correct
                if obs.isHuman == 'Y':
                    is_human_daily_counts[local_date] += 1  # Ignore interpreter warning, the if statement is correct
                if obs.isKong == 'N' and obs.isDog == 'N' and obs.isHuman == 'N':
                    no_dog_stance_counts[local_date] += 1  # Ignore interpreter warning, the if statement is correct

        # Count occurrences per day
        observation_counts = Counter(
            timezone.localtime(obs.obsDateTime, local_tz).date() for obs in observations
        )

        for year in range(first_date.year, last_date.year + 1):
            heatmap_data = []
            for single_date, count in observation_counts.items():
                if single_date.year == year:
                    day = single_date.day - 1
                    month = single_date.month - 1
                    kong_count = kong_daily_counts[single_date]
                    if branch == 'Italy':  # Italy branch has additional fields to prepare
                        dog_count = is_dog_daily_counts[single_date]
                        human_count = is_human_daily_counts[single_date]
                        no_stance_count = no_dog_stance_counts[single_date]
                        heatmap_data.append([day, month, count, kong_count, dog_count, human_count, no_stance_count])
                    else:
                        heatmap_data.append([day, month, count, kong_count])

            daily_heatmap_data[year] = heatmap_data

        # Iterate over the daily heatmap data to aggregate into weekly data
        for year, daily_data in daily_heatmap_data.items():
            if branch == 'Italy':  # Italy branch has additional fields to prepare
                for day, month, total_count, kong_count, dog_count, human_count, no_dog_stance_counts in daily_data:
                    # Calculate week number (1-5) within the month
                    week_number = math.ceil((day + 1) / 7) - 1
                    # Aggregate counts by week
                    weekly_key = (week_number, month)
                    weekly_heatmap_data[year][weekly_key][0] += total_count
                    weekly_heatmap_data[year][weekly_key][1] += kong_count
                    weekly_heatmap_data[year][weekly_key][2] += dog_count
                    weekly_heatmap_data[year][weekly_key][3] += human_count
                    weekly_heatmap_data[year][weekly_key][4] += no_dog_stance_counts
            else:
                for day, month, total_count, kong_count in daily_data:
                    # Calculate week number (1-5) within the month
                    week_number = math.ceil((day + 1) / 7) - 1
                    # Aggregate counts by week
                    weekly_key = (week_number, month)
                    weekly_heatmap_data[year][weekly_key][0] += total_count
                    weekly_heatmap_data[year][weekly_key][1] += kong_count

        # Convert to the required format [[week, month, total_count, kong_count], ...]
        formatted_weekly_heatmap_data = {
            year: [[week, month] + counts for (week, month), counts in data.items()]
            for year, data in weekly_heatmap_data.items()
        }

        # Initialize Treatment/Examination/Placement/Session(Observes) form when adding new entries
        treatment_form = TreatmentForm(request.POST or None)
        examination_form = EntranceExaminationForm(request.POST or None)
        placement_form = DogPlacementForm(request.POST or None, request=request)
        session_form = ObservesForm(request.POST or None, request=request)

        # Get page numbers for each table from request
        treatments_page_number = request.GET.get('treatments_page', 1)
        examinations_page_number = request.GET.get('examinations_page', 1)
        placements_page_number = request.GET.get('placements_page', 1)
        sessions_page_number = request.GET.get('sessions_page', 1)
        MAX_PER_PAGE = 6  # Limit entries per page

        # Create paginators for all tables
        treatments_paginator = Paginator(
            Treatment.objects.filter(dog=dog_record).order_by('-treatmentDate'), MAX_PER_PAGE)
        examinations_paginator = Paginator(
            EntranceExamination.objects.filter(dog=dog_record).order_by('-examinationDate'), MAX_PER_PAGE)
        placements_paginator = Paginator(
            DogPlacement.objects.filter(dog=dog_record).order_by('-entranceDate'), MAX_PER_PAGE)
        sessions_paginator = Paginator(
            Observes.objects.filter(dog=dog_record).prefetch_related('observation_set').order_by('-sessionDate'),
            MAX_PER_PAGE)

        # Get the relevant page
        treatments = treatments_paginator.get_page(treatments_page_number)
        examinations = examinations_paginator.get_page(examinations_page_number)
        placements = placements_paginator.get_page(placements_page_number)
        sessions = sessions_paginator.get_page(sessions_page_number)

        # Handle user submitting a new Treatment/Examination/Placement/Session form
        if request.method == "POST":
            # Ensure only one form is submitted
            form_type = request.POST.get('form_type')

            # Check if it's a Treatment form
            if form_type == 'treatment_form':
                if treatment_form.is_valid():
                    new_treatment = treatment_form.save(commit=False)
                    new_treatment.dog = dog_record
                    new_treatment.save()

                    # If this is an AJAX request, send back the new treatments data
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        treatments_data = Treatment.objects.filter(dog=dog_record).order_by('-treatmentDate')[
                                          :MAX_PER_PAGE]
                        data = {
                            'data': [render_to_string('_treatment_row.html',
                                                      {'treatment': treatment}) for treatment in treatments_data],
                            'pagination': render_to_string('_dog_record_pagination.html',
                                                           {'paginated_data': treatments,
                                                            'param_name': 'treatments_page'})
                        }
                        return JsonResponse(data)
                    else:
                        # Redirect back to the dog_record_view to see the new treatment.
                        return redirect('dogs_app:dog_record', pk=dog_record.pk)
                else:
                    errors = treatment_form.errors.as_json()
                    return JsonResponse({'status': 'fail', 'errors': errors}, status=400)

            # Check if it's an Examination form
            elif form_type == 'examination_form':
                if examination_form.is_valid():
                    new_examination = examination_form.save(commit=False)
                    new_examination.dog = dog_record
                    new_examination.save()

                    # If this is an AJAX request, send back the new Examination data
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        examinations_data = EntranceExamination.objects.filter(dog=dog_record).order_by(
                            '-examinationDate')[:MAX_PER_PAGE]
                        data = {
                            'data': [render_to_string('_examination_row.html',
                                                      {'examination': examination}) for examination in
                                     examinations_data],
                            'pagination': render_to_string('_dog_record_pagination.html',
                                                           {'paginated_data': examinations,
                                                            'param_name': 'examinations_page'})
                        }
                        return JsonResponse(data)
                    else:
                        # Redirect back to the dog_record_view to see the new treatment.
                        return redirect('dogs_app:dog_record', pk=dog_record.pk)
                else:
                    # Form is not valid
                    errors = examination_form.errors.as_json()
                    return JsonResponse({'status': 'fail', 'errors': errors}, status=400)

            # Check if it's a Placement form
            elif form_type == 'placement_form':
                try:
                    if placement_form.is_valid():
                        new_placement = placement_form.save(commit=False)
                        new_placement.dog = dog_record
                        new_placement.save()

                        # If this is an AJAX request, send back the new Placement data
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            placements_data = DogPlacement.objects.filter(dog=dog_record).order_by('-entranceDate')[
                                              :MAX_PER_PAGE]
                            data = {
                                'data': [render_to_string('_placement_row.html',
                                                          {'placement': placement}) for placement in placements_data],
                                'pagination': render_to_string('_dog_record_pagination.html',
                                                               {'paginated_data': placements,
                                                                'param_name': 'placements_page'})
                            }
                            return JsonResponse(data)
                        else:
                            # Redirect back to the dog_record_view to see the new placement.
                            return redirect('dogs_app:dog_record', pk=dog_record.pk)
                    else:
                        errors = placement_form.errors.as_json()
                        return JsonResponse({'status': 'fail', 'errors': errors}, status=400)
                except IntegrityError as e:
                    error_message = "Duplicate Kennel Placement, this dog already has this kennel on that date."
                    # Construct the error structure
                    errors_dict = {'__all__': [{'message': error_message}]}
                    # Convert the error dictionary to a JSON string
                    errors_json_string = json.dumps(errors_dict, cls=DjangoJSONEncoder)
                    return JsonResponse({'status': 'fail', 'errors': errors_json_string}, status=400)
            # Check if it's a Session (Observes) form
            elif form_type == 'session_form':
                try:
                    if session_form.is_valid():
                        new_session = session_form.save(commit=False)
                        new_session.dog = dog_record
                        new_session.save()

                        # If this is an AJAX request, send back the new Session data
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            sessions_data = Observes.objects.filter(dog=dog_record).order_by('-sessionDate')[:MAX_PER_PAGE]
                            data = {
                                'data': [render_to_string('_session_row.html',
                                                          {'session': session}) for session in sessions_data],
                                'pagination': render_to_string('_dog_record_pagination.html',
                                                               {'paginated_data': sessions,
                                                                'param_name': 'sessions_page'})
                            }
                            return JsonResponse(data)
                        else:
                            # Redirect back to the dog_record_view to see the new placement.
                            return redirect('dogs_app:dog_record', pk=dog_record.pk)
                    else:
                        errors = session_form.errors.as_json()
                        return JsonResponse({'status': 'fail', 'errors': errors}, status=400)
                except IntegrityError as e:
                    error_message = "Duplicate Camera Sessions, this dog already has this camera on that date."
                    # Construct the error structure
                    errors_dict = {'__all__': [{'message': error_message}]}
                    # Convert the error dictionary to a JSON string
                    errors_json_string = json.dumps(errors_dict, cls=DjangoJSONEncoder)
                    return JsonResponse({'status': 'fail', 'errors': errors_json_string}, status=400)

        # Check if request is AJAX call for switching pages
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            data = {
                'data': [],
                'pagination': ''
            }
            if 'treatments_page' in request.GET:
                data['data'] = [render_to_string('_treatment_row.html',
                                                 {'treatment': treatment}) for treatment in treatments]
                data['pagination'] = render_to_string('_dog_record_pagination.html',
                                                      {'paginated_data': treatments,
                                                       'param_name': 'treatments_page'})
            elif 'examinations_page' in request.GET:
                data['data'] = [render_to_string('_examination_row.html',
                                                 {'examination': examination}) for examination in examinations]
                data['pagination'] = render_to_string('_dog_record_pagination.html',
                                                      {'paginated_data': examinations,
                                                       'param_name': 'examinations_page'})
            elif 'placements_page' in request.GET:
                data['data'] = [render_to_string('_placement_row.html',
                                                 {'placement': placement}) for placement in placements]
                data['pagination'] = render_to_string('_dog_record_pagination.html',
                                                      {'paginated_data': placements,
                                                       'param_name': 'placements_page'})
            elif 'sessions_page' in request.GET:
                data['data'] = [render_to_string('_session_row.html',
                                                 {'session': session}) for session in sessions]
                data['pagination'] = render_to_string('_dog_record_pagination.html',
                                                      {'paginated_data': sessions,
                                                       'param_name': 'sessions_page'})
            return JsonResponse(data)

        # Generate Owner form
        owner = Owner.objects.filter(ownerSerialNum=dog_record.owner_id).first()
        owner_form = OwnerForm(request.POST or None, instance=owner)

        context = {
            'dog_record': dog_record,
            'owner': owner_form if owner_form else None,
            'treatments': treatments,
            'examinations': examinations,
            'placements': placements,
            'sessions': sessions,
            'treatment_form': treatment_form,
            'examination_form': examination_form,
            'placement_form': placement_form,
            'session_form': session_form,
            'daily_heatmap_data': json.dumps(daily_heatmap_data),
            'weekly_heatmap_data': json.dumps(formatted_weekly_heatmap_data),
            'heatmap_first_date': json.dumps(first_date.isoformat()),
            'heatmap_last_date': json.dumps(last_date.isoformat()),
            'heatmap_first_weekly_date': json.dumps(first_weekly_date.isoformat()),
        }
        return render(request, 'dog_record.html', context=context)

    # User is NOT logged in --> send them to login page
    else:
        messages.error(request, message='You Must Be Logged In To View That Page...')
        return redirect('dogs_app:home')


@vet_required
# Handle deleting a Treatment
def delete_treatment(request, treatment_id):
    if request.method == 'POST' and request.POST.get('_method') == 'DELETE' and request.headers.get(
            'X-Requested-With') == 'XMLHttpRequest':
        try:
            treatment = Treatment.objects.get(treatmentID=treatment_id)
            treatment.delete()
            return JsonResponse({'status': 'success'})
        except Treatment.DoesNotExist:
            return JsonResponse({'status': 'fail'})


@vet_required
# Handle deleting an Examination
def delete_examination(request, examination_id):
    if request.method == 'POST' and request.POST.get('_method') == 'DELETE' and request.headers.get(
            'X-Requested-With') == 'XMLHttpRequest':
        try:
            examination = EntranceExamination.objects.get(examinationID=examination_id)
            examination.delete()
            return JsonResponse({'status': 'success'})
        except EntranceExamination.DoesNotExist:
            return JsonResponse({'status': 'fail'})


@vet_required
# Handle deleting a DogPlacement
def delete_placement(request, placement_id):
    if request.method == 'POST' and request.POST.get('_method') == 'DELETE' and request.headers.get(
            'X-Requested-With') == 'XMLHttpRequest':
        try:
            placement = DogPlacement.objects.get(id=placement_id)
            placement.delete()
            return JsonResponse({'status': 'success'})
        except DogPlacement.DoesNotExist:
            return JsonResponse({'status': 'fail'})


@vet_required
# Handle deleting a Session (Observes)
def delete_session(request, session_id):
    if request.method == 'POST' and request.POST.get('_method') == 'DELETE' and request.headers.get(
            'X-Requested-With') == 'XMLHttpRequest':
        try:
            session = Observes.objects.get(id=session_id)
            session.delete()
            return JsonResponse({'status': 'success'})
        except Observes.DoesNotExist:
            return JsonResponse({'status': 'fail'})


@vet_required
# Handle editing a Treatment
def edit_treatment(request, treatment_id):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.method == 'POST' and request.user.is_authenticated:
            try:
                treatment = Treatment.objects.get(treatmentID=treatment_id)
                form = TreatmentForm(request.POST, instance=treatment)
                if form.is_valid():
                    form.save()
                    return JsonResponse({'status': 'success'})
                else:
                    errors = {}
                    for key, value in form.errors.items():
                        errors[key] = ', '.join([str(error) for error in value])
                    return JsonResponse({'status': 'fail', 'errors': errors})
            except Treatment.DoesNotExist:
                return JsonResponse({'status': 'fail'})
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        elif request.method == 'GET':
            try:
                treatment = Treatment.objects.get(treatmentID=treatment_id)
                treatment_data = {
                    'status': 'success',
                    'treatmentName': treatment.treatmentName if treatment.treatmentName else None,
                    'treatmentDate': treatment.treatmentDate.isoformat() if treatment.treatmentDate else None,
                    'treatedBy': treatment.treatedBy if treatment.treatedBy else None,
                    'comments': treatment.comments if treatment.comments else None,
                }
                return JsonResponse(treatment_data)
            except Treatment.DoesNotExist:
                return JsonResponse({'status': 'fail', 'message': 'Treatment not found'})
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        else:
            # Return a Method Not Allowed status if not GET or POST
            return HttpResponseNotAllowed(['POST', 'GET'])


@vet_required
# Handle editing an Examination
def edit_examination(request, examination_id):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.method == 'POST' and request.user.is_authenticated:
            try:
                examination = EntranceExamination.objects.get(examinationID=examination_id)
                form = EntranceExaminationForm(request.POST, instance=examination)
                if form.is_valid():
                    form.save()
                    return JsonResponse({'status': 'success'})
                else:
                    errors = {}
                    for key, value in form.errors.items():
                        errors[key] = ', '.join([str(error) for error in value])
                    return JsonResponse({'status': 'fail', 'errors': errors})
            except EntranceExamination.DoesNotExist:
                return JsonResponse({'status': 'fail'})
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        elif request.method == 'GET':
            try:
                examination = EntranceExamination.objects.get(examinationID=examination_id)
                examination_data = {
                    'status': 'success',
                    'examinationDate': examination.examinationDate.isoformat() if examination.examinationDate else None,
                    'examinedBy': examination.examinedBy if examination.examinedBy else None,
                    'results': examination.results if examination.results else None,
                    'dogWeight': examination.dogWeight if examination.dogWeight else None,
                    'dogTemperature': examination.dogTemperature if examination.dogTemperature else None,
                    'dogPulse': examination.dogPulse if examination.dogPulse else None,
                    'comments': examination.comments if examination.comments else None,
                }
                return JsonResponse(examination_data)
            except EntranceExamination.DoesNotExist:
                return JsonResponse({'status': 'fail', 'message': 'Examination not found'})
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        else:
            # Return a Method Not Allowed status if not GET or POST
            return HttpResponseNotAllowed(['POST', 'GET'])


@vet_required
# Handle editing a DogPlacement
def edit_placement(request, placement_id):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.method == 'POST' and request.user.is_authenticated:
            try:
                placement = DogPlacement.objects.get(id=placement_id)
                form = DogPlacementForm(request.POST, instance=placement, request=request)
                if form.is_valid():
                    form.save()
                    return JsonResponse({'status': 'success'})
                else:
                    errors = {}
                    for key, value in form.errors.items():
                        errors[key] = ', '.join([str(error) for error in value])
                    return JsonResponse({'status': 'fail', 'errors': errors})
            except DogPlacement.DoesNotExist:
                return JsonResponse({'status': 'fail'})
            # Handle IntegrityError for duplicate kennel placements
            except IntegrityError:
                error_message = 'Duplicate Kennel Placement, this dog already has this kennel on that date.'
                return JsonResponse({'status': 'fail', 'errors': {'__all__': [error_message]}})
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        elif request.method == 'GET':
            try:
                placement = DogPlacement.objects.get(id=placement_id)
                placement_data = {
                    'status': 'success',
                    'kennel': serializers.serialize('json',
                                                    [placement.kennel]) if placement.kennel else None,
                    'entranceDate': placement.entranceDate.isoformat() if placement.entranceDate else None,
                    'expirationDate': placement.expirationDate.isoformat() if placement.expirationDate else None,
                    'placementReason': placement.placementReason if placement.placementReason else None,
                }
                return JsonResponse(placement_data)
            except DogPlacement.DoesNotExist:
                return JsonResponse({'status': 'fail', 'message': 'Placement not found'})
                # Handle IntegrityError for duplicate kennel placements
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        else:
            # Return a Method Not Allowed status if not GET or POST
            return HttpResponseNotAllowed(['POST', 'GET'])


@vet_required
# Handle editing a Session (Observes)
def edit_session(request, session_id):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.method == 'POST' and request.user.is_authenticated:
            try:
                session = Observes.objects.get(id=session_id)
                form = ObservesForm(request.POST, instance=session, request=request)
                if form.is_valid():
                    form.save()
                    return JsonResponse({'status': 'success'})
                else:
                    errors = {}
                    for key, value in form.errors.items():
                        errors[key] = ', '.join([str(error) for error in value])
                    return JsonResponse({'status': 'fail', 'errors': errors})
            except Observes.DoesNotExist:
                return JsonResponse({'status': 'fail'})
            except IntegrityError:
                error_message = 'Duplicate Camera Session, this dog already has this camera on that date.'
                return JsonResponse({'status': 'fail', 'errors': {'__all__': [error_message]}})
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        elif request.method == 'GET':
            try:
                session = Observes.objects.get(id=session_id)
                session_data = {
                    'status': 'success',
                    'camera': serializers.serialize('json', [session.camera]) if session.camera else None,
                    'sessionDate': session.sessionDate.isoformat() if session.sessionDate else None,
                    'comments': session.comments if session.comments else None,
                }
                return JsonResponse(session_data)
            except Observes.DoesNotExist:
                return JsonResponse({'status': 'fail', 'message': 'Session not found'})
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        else:
            # Return a Method Not Allowed status if not GET or POST
            return HttpResponseNotAllowed(['POST', 'GET'])


@vet_required
# Handle Editing an Owner
def edit_owner(request, owner_id):
    # Get the current Branch
    branch = get_current_branch(request)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.method == 'POST' and request.user.is_authenticated:
            try:
                owner = Owner.objects.get(ownerSerialNum=owner_id)

                # Validate that the user is in the correct branch
                if owner.branch != branch:
                    messages.error(request, "You must be in the correct branch to edit this owner...")
                    return redirect('dogs_app:home')

                form = OwnerForm(request.POST, instance=owner)
                if form.is_valid():
                    form.save()
                    return JsonResponse({'status': 'success'})
                else:
                    errors = {}
                    for key, value in form.errors.items():
                        errors[key] = ', '.join([str(error) for error in value])
                    return JsonResponse({'status': 'fail', 'errors': errors})
            except Owner.DoesNotExist:
                return JsonResponse({'status': 'fail'})
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        elif request.method == 'GET':
            try:
                owner = Owner.objects.get(ownerSerialNum=owner_id)

                # Validate that the user is in the correct branch
                if owner.branch != branch:
                    messages.error(request, "You must be in the correct branch to edit this owner...")
                    return redirect('dogs_app:home')

                owner_data = {
                    'status': 'success',
                    'firstName': owner.firstName if owner.firstName else None,
                    'lastName': owner.lastName if owner.lastName else None,
                    'ownerID': owner.ownerID if owner.ownerID else None,
                    'ownerAddress': owner.ownerAddress if owner.ownerAddress else None,
                    'city': owner.city if owner.city else None,
                    'phoneNum': owner.phoneNum if owner.phoneNum else None,
                    'cellphoneNum': owner.cellphoneNum if owner.cellphoneNum else None,
                    'comments': owner.comments if owner.comments else None,
                }
                return JsonResponse(owner_data)
            except Owner.DoesNotExist:
                return JsonResponse({'status': 'fail', 'message': 'Owner not found'})
            except Exception as e:
                return JsonResponse({'status': 'fail', 'message': str(e)})
        else:
            # Return a Method Not Allowed status if not GET or POST
            return HttpResponseNotAllowed(['POST', 'GET'])


@vet_required
# Handle Observations display
def view_observations(request, session_id):
    observation_form = None
    stance_form = None
    if request.user.is_authenticated:
        dog_stances = DogStance.objects.filter(observation__observes_id=session_id).order_by('stanceStartTime')
        observations = Observation.objects.filter(observes_id=session_id).prefetch_related(
            Prefetch('dogstance_set', queryset=dog_stances, to_attr='related_dog_stances')
        ).order_by('-obsDateTime')
        session_instance = Observes.objects.get(pk=session_id)

        # Check if user is on the correct branch
        branch = get_current_branch(request)
        dog = session_instance.dog
        if dog.branch != branch:
            messages.error(request, "You must be in the correct branch to view this dog's observations...")
            return redirect('dogs_app:home')

        elif request.method == 'POST':
            # Check the form type to determine which form is being submitted
            form_type = request.POST.get('form_type')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and form_type == 'dog_stance':
                # Handle DogStance form submission via AJAX
                stance_form = DogStanceForm(request.POST or None)
                if stance_form.is_valid():
                    try:
                        new_stance = stance_form.save(commit=False)
                        new_stance.observation_id = request.POST.get('observation_id')
                        new_stance.save()
                        messages.success(request, 'Stance has been added successfully!')
                        return JsonResponse({
                            "status": "success",
                            "new_stance": {
                                'id': new_stance.id,
                                'stanceStartTime': new_stance.stanceStartTime,
                                'dogStance': new_stance.get_dogStance_display(),
                                'dogLocation': new_stance.get_dogLocation_display() if new_stance.dogLocation else "-",
                                'observation': new_stance.observation_id,
                            }
                        }, status=201)
                    except ValidationError as e:
                        return JsonResponse({"status": "error", "errors": e.message_dict}, status=400)
                    except IntegrityError:
                        return JsonResponse({"status": "error",
                                             "errors": "Duplicate Stance Start Time"}, status=400)
                    except Exception as e:
                        return JsonResponse({"status": "error", "errors": str(e)}, status=400)
                else:
                    return JsonResponse({"status": "error", "errors": stance_form.errors}, status=400)
            elif request.headers.get('X-Requested-With') == 'XMLHttpRequest' and form_type == 'add_observation':
                # Handle Observation form submission via AJAX
                observation_form = ObservationForm(request.POST or None, request.FILES or None, request=request)
                if observation_form.is_valid():
                    try:
                        new_observation = observation_form.save(commit=False)
                        new_observation.observes = session_instance

                        # Convert to naive datetime
                        naive_dt = new_observation.obsDateTime.replace(tzinfo=None)
                        # Apply the correct timezone
                        local_tz = pytz.timezone('Asia/Jerusalem')
                        local_dt = local_tz.localize(naive_dt)
                        # Convert to UTC
                        utc_dt = local_dt.astimezone(pytz.utc)
                        new_observation.obsDateTime = utc_dt

                        # If user is not in Italy branch, set the isDog and isHuman fields to None
                        if branch.branchName != 'Italy':
                            new_observation.isDog = None
                            new_observation.isHuman = None

                        new_observation.save()
                        messages.success(request, 'Observation has been added successfully!')
                        return JsonResponse({"status": "success"}, status=201)
                    except ValidationError as e:
                        return JsonResponse({"status": "error", "errors": e.message_dict}, status=400)
                    except IntegrityError:
                        return JsonResponse({"status": "error", "errors": "Duplicate Observation Date"}, status=400)
                    except Exception as e:
                        return JsonResponse({"status": "error", "errors": str(e)}, status=400)
                else:
                    return JsonResponse({"status": "error", "errors": observation_form.errors}, status=400)
        else:
            observation_form = ObservationForm(request=request)
            stance_form = DogStanceForm()

        # Pagination
        paginator = Paginator(observations, ENTRIES_PER_PAGE)
        page_number = request.GET.get('page')
        paginated_observations = paginator.get_page(page_number)
        context = {
            'observations': observations,
            'paginated_observations': paginated_observations,
            'session_instance': session_instance,
            'observation_form': observation_form if observation_form else None,
            'stance_form': stance_form if stance_form else None,
        }

        return render(request, 'view_observations.html', context=context)
    else:
        messages.error(request, message='You Must Be Logged In To View That Page...')
        return redirect('dogs_app:home')


@vet_required
# Handle deleting an Observation
@require_POST  # Ensures this view can only be accessed with POST request
def delete_observation(request):
    if request.method == 'POST' and request.user.is_authenticated and request.headers.get(
            'X-Requested-With') == 'XMLHttpRequest':
        observation_id = request.POST.get('observation_id')

        try:
            session_id = Observes.objects.get(observation__id=observation_id).id
            dog_stances = DogStance.objects.filter(observation_id=observation_id).order_by('stanceStartTime')
            observations = Observation.objects.filter(observes_id=session_id).prefetch_related(
                Prefetch('dogstance_set', queryset=dog_stances, to_attr='related_dog_stances')
            ).order_by('-obsDateTime')
        except Observes.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Session not found"}, status=404)
        # Get the current page number
        page_number = request.POST.get('page_number', 1)  # Get the current page number
        if page_number == 'null':
            page_number = 1
        paginator = Paginator(observations, ENTRIES_PER_PAGE)

        # Check if the current page is empty
        try:
            current_page = paginator.page(page_number)
            is_current_page_empty = len(current_page.object_list)-1 == 0
        except EmptyPage:
            is_current_page_empty = True

        try:
            observation = Observation.objects.get(id=observation_id)
            observation.delete()

            return JsonResponse({"status": "success",
                                 "is_current_page_empty": is_current_page_empty
                                 })
        except Observation.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Observation not found"}, status=404)
    else:
        return JsonResponse({"status": "error", "message": "Unauthorized"}, status=401)


@vet_required
# Handle editing an Observation
def edit_observation(request, observation_id):
    if request.method == 'GET':
        observation = Observation.objects.get(id=observation_id)
        observation_form = ObservationForm(instance=observation, request=request)

        # Exclude the file field from the JSON response
        observation_data = observation_form.initial

        # Include the original CSV file name and the URL if available
        if observation.csvFile:
            observation_data['original_csv_file_name'] = observation.original_csv_file_name if observation.original_csv_file_name else observation.csvFile.name
            observation_data['csvFile'] = observation.csvFile.url
        else:
            observation_data['original_csv_file_name'] = None
            observation_data['csvFile'] = None

        # Include the original Video file name and the URL if available
        if observation.rawVideo:
            observation_data['original_video_file_name'] = observation.original_video_file_name if observation.original_video_file_name else observation.rawVideo.name
            observation_data['rawVideo'] = observation.rawVideo.url
        else:
            observation_data['original_video_file_name'] = None
            observation_data['rawVideo'] = None

        return JsonResponse({"status": "success", "observation": observation_data}, status=200)

    elif request.method == 'POST':
        branch = get_current_branch(request)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            observation_form = ObservationForm(request.POST or None, request.FILES or None,
                                               instance=Observation.objects.get(id=observation_id), request=request)
            if observation_form.is_valid():
                try:
                    new_observation = observation_form.save(commit=False)

                    # Convert to naive datetime
                    naive_dt = new_observation.obsDateTime.replace(tzinfo=None)
                    # Apply the correct timezone
                    local_tz = pytz.timezone('Asia/Jerusalem')
                    local_dt = local_tz.localize(naive_dt)
                    # Convert to UTC
                    utc_dt = local_dt.astimezone(pytz.utc)
                    new_observation.obsDateTime = utc_dt

                    # If user is not in Italy branch, make sure the isDog and isHuman fields are None
                    if branch.branchName != 'Italy':
                        new_observation.isDog = None
                        new_observation.isHuman = None

                    new_observation.save()

                    # Get the updated observation
                    updated_observation = Observation.objects.get(id=observation_id)

                    # Render the _observation_row.html template with the updated observation
                    new_row_html = render_to_string('_observation_row.html', {'observation': updated_observation}, request=request)

                    # Create a serializable observation data dictionary
                    observation_data = observation_form.cleaned_data
                    observation_data['csvFile'] = updated_observation.csvFile.url if updated_observation.csvFile else None
                    observation_data['rawVideo'] = updated_observation.rawVideo.url if updated_observation.rawVideo else None
                    observation_data['original_csv_file_name'] = updated_observation.original_csv_file_name if updated_observation.original_csv_file_name else updated_observation.csvFile.name
                    observation_data['original_video_file_name'] = updated_observation.original_video_file_name if updated_observation.original_video_file_name else updated_observation.rawVideo.name


                    messages.success(request, 'Observation has been edited successfully!')
                    return JsonResponse(
                        {"status": "success", "observation": observation_form.cleaned_data, "newRowHtml": new_row_html},
                        status=200)
                except ValidationError as e:
                    return JsonResponse({"status": "error", "errors": e.message_dict}, status=400)
                except IntegrityError as e:
                    return JsonResponse({"status": "error", "errors": "Duplicate Observation Date"}, status=400)
                except Exception as e:
                    return JsonResponse({"status": "error", "errors": str(e)}, status=400)
            else:
                return JsonResponse({"status": "error", "errors": observation_form.errors}, status=400)
        else:
            return HttpResponseNotAllowed(['POST', 'GET'])


@vet_required
# Handle deleting a DogStance
@require_POST  # Ensures this view can only be accessed with POST request
def delete_stance(request):
    if request.method == 'POST' and request.user.is_authenticated and request.headers.get(
            'X-Requested-With') == 'XMLHttpRequest':
        stance_id = request.POST.get('stance_id')
        try:
            dog_stance = DogStance.objects.get(id=stance_id)
            dog_stance.delete()
            return JsonResponse({"status": "success"})
        except DogStance.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Dog Stance not found"}, status=404)
    else:
        return JsonResponse({"status": "error", "message": "Unauthorized"}, status=401)


@vet_required
# Handle editing a DogStance
def edit_dog_stance(request, stance_id):
    if request.method == 'GET':
        try:
            stance = DogStance.objects.get(id=stance_id)
            stance_form = DogStanceForm(instance=stance)
            stance_data = stance_form.initial
            return JsonResponse({"status": "success", "stance": stance_data}, status=200)
        except DogStance.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Dog Stance not found"}, status=404)
    elif request.method == 'POST':
        stance_form = DogStanceForm(request.POST or None, instance=DogStance.objects.get(id=stance_id))
        if stance_form.is_valid():
            try:
                saved_stance = stance_form.save()
                observation_id = saved_stance.observation.id
                observation = Observation.objects.get(id=observation_id)
                updated_stance = DogStance.objects.get(id=stance_id)
                new_row_html = render_to_string('_observation_row.html', {'stance': updated_stance, 'observation': observation}, request=request)
                messages.success(request, 'Stance has been edited successfully!')
                return JsonResponse(
                    {"status": "success",
                     "stance": stance_form.cleaned_data,
                     "newRowHtml": new_row_html,
                     "observationId": observation_id},
                    status=200)
            except ValidationError as e:
                return JsonResponse({"status": "error", "errors": e.message_dict}, status=400)
            except IntegrityError:
                return JsonResponse({"status": "error", "errors": "Duplicate Stance Start Time"}, status=400)
            except Exception as e:
                return JsonResponse({"status": "error", "errors": str(e)}, status=400)
        else:
            return JsonResponse({"status": "error", "errors": stance_form.errors}, status=400)
    else:
        return HttpResponseNotAllowed(['POST', 'GET'])


@vet_required
# Deleting a dog record
def delete_dog_view(request, pk):
    # Check if the user is logged in
    if request.user.is_authenticated:
        # Check if the user has the right permissions
        if not request.user.has_perm('dogs_app.delete_dog'):
            raise PermissionDenied
        # Grab the dog to delete
        delete_dog = get_object_or_404(Dog, dogID=pk)

        if request.method == 'POST':
            dog_name = delete_dog.dogName
            delete_dog.delete()
            messages.success(request, f'{dog_name} Has Been Deleted Successfully...')
            return redirect('dogs_app:view_dogs')
        # User clicked "Delete" button to confirm deletion
        return render(request, 'delete_dog.html', {'dog': delete_dog})
    # User not logged in, must login first
    else:
        messages.error(request, 'You must be logged in to do that...')
        return redirect('dogs_app:home')


@vet_required
def add_dog_view(request):
    branch = get_current_branch(request)

    # Check if user is logged in
    if request.user.is_authenticated:
        # Check if the user has the right permissions
        if not request.user.has_perm('dogs_app.add_dog'):
            raise PermissionDenied

        # If form is submitted (i.e., User has filled the form)
        if request.method == 'POST':
            # Initialize the form
            form = AddDogForm(request.POST, request.FILES, request=request)
            # Validate the form inputs
            if form.is_valid():
                try:
                    # Create Dog instance but don't save to DB yet
                    dog_instance = form.save(commit=False)
                    # Assign the current branch to the Dog
                    dog_instance.branch = branch
                    # Save the Dog instance to the database
                    dog_instance.save()
                    messages.success(request, f"{form.cleaned_data['dogName']} Has Been Added Successfully...")
                except Exception as e:
                    messages.error(request, f"Error occurred while adding Dog: {str(e)}")

                return redirect('dogs_app:view_dogs')
            # If form is not valid, render errors
            else:
                return render(request, 'add_dog.html', {"form": form})
        # If request is not POST (i.e., GET), just render the form
        else:
            form = AddDogForm(request=request)
            return render(request, 'add_dog.html', {"form": form})
    # If user is not logged in/authenticated, show an error message and redirect to home.
    else:
        messages.error(request, "You Must Be Logged In To Add New Dogs...")
        return redirect('dogs_app:home')


@vet_required
# Updating a Dog record
def update_dog_view(request, pk):
    # Check if user is logged in
    if request.user.is_authenticated:
        # Check if the user has the right permissions to edit dogs
        if not request.user.has_perm('dogs_app.change_dog'):
            raise PermissionDenied

        # Grab the Dog record
        current_dog = Dog.objects.get(dogID=pk)
        form = AddDogForm(request.POST or None, request.FILES or None, instance=current_dog, request=request)

        # Check if the "deleteImage" button was clicked
        if 'deleteImage' in request.POST:
            # Delete the dog's image if it is not the default images
            if current_dog.dogImage and current_dog.dogImage.name:
                os.remove(os.path.join(settings.MEDIA_ROOT, current_dog.dogImage.name))

            current_dog.dogImage = None
            current_dog.save()
            messages.success(request, f"{current_dog.dogName}'s Picture Has Been Removed Successfully!")
            # Refresh the current page
            return HttpResponseRedirect(request.path_info)

        # Check if form is submitted
        if form.is_valid():
            form.save()
            messages.success(request, f"{current_dog.dogName}'s Details Have Been Updated Successfully!")
            return redirect('dogs_app:dog_record', pk=current_dog.dogID)
        return render(request, 'update_dog.html', {'form': form, 'current_dogID': current_dog.dogID})

    # User is not logged in, redirect them to login
    else:
        messages.error(request, "You must be logged in to do that...")
        return redirect('dogs_app:home')


@superuser_required
# Display all users for the admins only to view
def view_users(request):
    # Retrieve sorting criteria from request
    order_by = request.GET.get('order_by', 'username')  # Default sort field
    direction = request.GET.get('direction', 'asc')  # Default sort direction

    # Retrieve all the users in the system, prefetch their groups, and select their profiles
    users = User.objects.all().prefetch_related('groups').select_related('profile')

    # Use a Role Filter for filtering Users based on their Role
    role_filter = request.GET.get('role', 'all')

    page = request.GET.get('page')

    # Apply role filter
    if role_filter and role_filter != "all":
        if role_filter == "Admin":
            users = users.filter(is_superuser=True)
        else:
            users = users.filter(groups__name=role_filter)

    # Search Query
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(Q(username__icontains=search_query) | Q(email__icontains=search_query) | Q(
            first_name__icontains=search_query) | Q(last_name__icontains=search_query))

    # Apply sorting
    if order_by == 'full_name':
        users = users.annotate(full_name=Concat('first_name', Value(' '), 'last_name'))
    elif order_by in ['phone_number', 'address']:
        order_by = 'profile__' + order_by

    if order_by == 'role':
        users_page, paginator = sort_users_by_role(users, direction, page)
    else:
        if direction == 'desc':
            sort_order = '-' + order_by
        else:
            sort_order = order_by
        users = users.order_by(sort_order)

        # Create a Paginator object with the users and the number of users per page
        paginator = Paginator(users, ENTRIES_PER_PAGE, allow_empty_first_page=True)

        # Get the current page number from the request's GET parameters
        page = request.GET.get('page')

        # Get the Page object for the current page
        users_page = paginator.get_page(page)

    # Calculate pagination range
    pagination_start = max(users_page.number - 3, 1)
    pagination_end = min(users_page.number + 3, paginator.num_pages)

    # Fetch each user's Status (E.g: "Admin", "Vet" or "Viewer")
    for user in users_page:
        user.role = get_user_role(user)

    context = {
        'users': users_page,
        'pagination_start': pagination_start,
        'pagination_end': pagination_end,
        'role_filter': role_filter,
        'order_by': order_by,
        'direction': direction,
        'search_query': search_query,
    }

    # pass all users on to viewers_users.html
    return render(request, 'view_users.html', context=context)


# Takes in a User, returns their ranking.
# Used in view_users and update_user to determine User Status
def get_user_role(user):
    if user.is_superuser:
        return "Admin"
    elif user.groups.filter(name="Vet").exists():
        return "Vet"
    elif user.groups.filter(name="Viewer").exists():
        return "Viewer"
    else:
        return ""


# Custom sorting function for view_users
def sort_users_by_role(users, direction='asc', page_number=1):
    # Mapping roles to sort order
    ROLE_SORT_ORDER = {'Admin': 0, 'Vet': 1, 'Viewer': 2}

    # Use the existing get_user_role function to extract the role's sort order
    def get_role_order(user):
        role = get_user_role(user)
        return ROLE_SORT_ORDER.get(role, -1)

    # Sort users based on the role order and username
    sorted_users = sorted(users, key=lambda user: (get_role_order(user), user.username))

    # If the direction is descending, reverse the sorted list
    if direction == 'desc':
        sorted_users.reverse()

    # Create a Paginator object with the sorted_users and the number of users per page
    paginator = Paginator(sorted_users, ENTRIES_PER_PAGE, allow_empty_first_page=True)

    # Get the Page object for the current page
    users_page = paginator.get_page(page_number)

    return users_page, paginator


@superuser_required
# Delete a user in the Admin user-view page
def delete_user_view(request, pk):
    # Check if the user is logged in
    if request.user.is_authenticated:
        # Check if the user has the right permissions
        if not request.user.has_perm('dogs_app.delete_user'):
            raise PermissionDenied
        # Grab the user to delete
        delete_user = get_object_or_404(User, pk=pk)
        # User trying to display the delete page
        if request.method == 'POST':
            user_name = delete_user.username
            delete_user.delete()
            messages.success(request, f'{user_name} Has Been Deleted Successfully...')
            return redirect('dogs_app:view_users')
        # User clicked "Delete" button to confirm deletion
        return render(request, 'delete_user.html', {'delete_user': delete_user})
    # User not logged in, must login first
    else:
        messages.error(request, message='You must be logged in to do that...')
        return redirect('dogs_app:home')


@superuser_required
# View for updating user details for admins
def update_user_view(request, pk):
    # Check if user is logged in
    if request.user.is_authenticated:
        user_to_update = User.objects.get(pk=pk)
        initial = {'role': get_user_role(user_to_update)}
        # Main User Form
        user_form = UpdateUserForm(request.POST or None,
                                   instance=user_to_update,
                                   initial=initial,
                                   request_user=request.user)
        # Addition User Details Form (phone, address, image)
        profile_form = ProfileUpdateForm(request.POST or None,
                                         request.FILES or None,
                                         instance=user_to_update.profile)

        # Check if the "deleteImage" button was clicked
        if 'deleteImage' in request.POST:
            # Delete the user's profile image if it is not the default image
            if user_to_update.profile.image and 'default.jpg' not in user_to_update.profile.image.name:
                os.remove(os.path.join(settings.MEDIA_ROOT, user_to_update.profile.image.name))

            user_to_update.profile.image = DEFAULT_IMAGE_SOURCE
            user_to_update.profile.save()
            messages.success(request, f"{user_to_update.first_name} {user_to_update.last_name}'s"
                                      f" Picture Has Been Removed Successfully!")
            # Refresh the current page
            return HttpResponseRedirect(request.path_info)

        if user_form.is_valid() and profile_form.is_valid():
            # Save form without committing (we'll modify the user before saving)
            user = user_form.save(commit=False)
            role = user_form.cleaned_data['role']
            # If the group name is not empty
            if role:
                # If group name is Admin, update is_superuser status and clear other groups
                if role == 'Admin':
                    user.is_superuser = True
                    user.is_staff = True
                    user_to_update.groups.clear()
                else:
                    # For other groups, get the chosen group
                    chosen_group = Group.objects.get(name=role)
                    # Remove the user from all the other groups
                    user_to_update.groups.clear()
                    # Add the user to the chosen group
                    user_to_update.groups.add(chosen_group)
                    # Update user's is_superuser status
                    user.is_superuser = False
                    user.is_staff = False

            # Now commit the save
            user.save()
            profile_form.save()
            messages.success(request, f"{user_to_update.first_name} {user_to_update.last_name}'s Details"
                                      f" Have Been Updated Successfully!")
            return redirect('dogs_app:view_users')
        context = {
            'user_form': user_form,
            'profile_form': profile_form,
            'user_to_update': user_to_update,
            'profile': user_to_update.profile
        }
        return render(request, 'update_user.html', context=context)
    # User is not logged in, redirect them to login
    else:
        messages.error(request, "You must be logged in to do that...")
        return redirect('dogs_app:home')


@login_required
# User updating their own details view
# Only logged-in users permitted
def update_user_self_view(request):
    if request.user.is_authenticated:
        user_to_update = User.objects.get(pk=request.user.pk)
        initial = {'role': get_user_role(user_to_update)}
        # Main User Form
        user_form = UpdateUserForm(request.POST or None,
                                   instance=user_to_update,
                                   initial=initial,
                                   request_user=user_to_update)

        profile_to_update = user_to_update.profile
        profile_form = ProfileUpdateForm(request.POST or None,
                                         request.FILES or None,
                                         instance=user_to_update.profile)
        if request.method == 'POST':
            if user_form.is_valid() and profile_form.is_valid():
                user = user_form.save(commit=False)
                # Ensure non-admin users can't change their own roles
                if not request.user.is_superuser and 'role' in user_form.changed_data:
                    messages.error(request, "You are not allowed to change your role.")
                    return render(request, 'update_user.html', {
                        'user_form': user_form,
                        'profile_form': profile_form,
                        'profile': profile_to_update,
                    })

                # Check if the delete image button was clicked
                if 'deleteImage' in request.POST:
                    # Delete the user's profile image if it is not the default image
                    if profile_to_update.image and 'default.jpg' not in profile_to_update.image.name:
                        os.remove(os.path.join(settings.MEDIA_ROOT, profile_to_update.image.name))

                    profile_to_update.image = DEFAULT_IMAGE_SOURCE
                    profile_to_update.save()

                    messages.success(request, "Your Picture Has Been Removed Successfully!")
                    # Refresh the current page
                    return HttpResponseRedirect(request.path_info)

                profile_form.save()
                user.save()

                messages.success(request, f"Your Details Have Been Updated Successfully!")
                return redirect('dogs_app:update_user_self')
            else:
                return render(request, 'update_user.html', {
                    'user_form': user_form,
                    'profile_form': profile_form,
                    'profile': profile_to_update,
                })
        else:
            context = {
                'user_form': user_form,
                'profile_form': profile_form,
                'profile_to_update': profile_to_update,
                'user_to_update': user_to_update
            }
            return render(request, 'update_user.html', context)
    else:
        messages.error(request, "You must be logged in to do that...")
        return redirect('dogs_app:home')


@superuser_required
# Page for editing the news from the homepage. Only visible to Admins
def update_news(request, news_id):
    # Get the news story to edit
    news = get_object_or_404(News, pk=news_id)

    # Get the current Branch
    branch = get_current_branch(request)

    if news.branch != branch:
        messages.error(request, "You must be in the correct branch to edit this news story...")
        return redirect('dogs_app:home')

    elif request.method == 'POST':
        news.title = request.POST['title']
        news.content = request.POST['content']
        news.save()
        messages.success(request, f"News Story: '{request.POST['title']}' has been successfully edited...")
        return redirect('dogs_app:home')
    else:
        form = NewsForm(instance=news)
        return render(request, 'update_news.html', {'news': news, 'form': form})


@superuser_required
# View for deleting a News story from the homepage. Only available to Admins
def delete_news(request, news_id):
    # Get the news story to delete
    news = get_object_or_404(News, pk=news_id)

    # Get the current Branch
    branch = get_current_branch(request)

    if news.branch != branch:
        messages.error(request, "You must be in the correct branch to delete this news story...")
        return redirect('dogs_app:home')

    elif request.method == 'POST':
        news_title = news.title
        news.delete()
        messages.success(request, f"News Story: '{news_title}' has been successfully deleted...")
        return redirect('dogs_app:home')
    else:
        return render(request, 'delete_news.html', {'news': news})


# Render Graphs page
def graphs(request):
    # Get the maximum number of stance options used in the DogStance in the website
    # For the slider max limit above the "Top Dog Stances Across The Week" chart
    max_dog_stances_count = get_max_dog_stances_count(request=request)
    return render(request, 'graphs.html', {'max_dog_stances_count': max_dog_stances_count})


# View for the Dynamic Graphs page
def chart_data(request):
    branch = get_current_branch(request)

    # Get all the Dogs that have received a Kong Toy
    dogs_with_kong = Dog.objects.filter(branch=branch, kongDateAdded__isnull=False)

    # ---Health Metrics Profiling of Our Canines Chart---
    # Prepare a dictionary in JSON for distribution of dogs by gender, vaccination and isneutered
    # ONLY for vets
    if request.user.is_authenticated and (request.user.groups.filter(name='Vet').exists() or request.user.is_superuser):
        health_metrics = get_health_metrics_dict(request=request)
    # -----------------------------------

    # ---Dog Stances With/Without Kong Charts---
    # Get all Observations with and without a Kong Toy
    obs_with_kong = Observation.objects.filter(observes__dog__branch=branch, isKong='Y', obsDateTime__isnull=False)
    obs_without_kong = Observation.objects.filter(observes__dog__branch=branch, isKong='N', obsDateTime__isnull=False)

    # Count of DogStances with and without kong toy, then make sure we use the front-end names of the values
    stances_with_kong = DogStance.objects.filter(observation__in=obs_with_kong).values('dogStance').annotate(
        total=models.Count('observation'))
    stances_with_kong = map_stances_to_frontend(list(stances_with_kong))
    stances_without_kong = DogStance.objects.filter(observation__in=obs_without_kong).values('dogStance').annotate(
        total=models.Count('observation'))
    stances_without_kong = map_stances_to_frontend(list(stances_without_kong))

    # Get a dictionary with count of dogStances by years for the "Dog Stances With/Without Kong" Charts yearly options
    yearly_stances_with_kong, yearly_stances_without_kong = get_stances_count_by_year(request=request)

    # For the drop-down yearly selection
    years_with_kong = sorted(yearly_stances_with_kong.keys(), reverse=True)
    years_without_kong = sorted(yearly_stances_without_kong.keys(), reverse=True)
    # -----------------------------------------

    # ---Most Common General Behaviors Chart---
    # Fetch a dictionary of the top combined stance+position in DogStances,
    # with their counts of "with" and "without" kong individually
    top_stance_position_combos = fetch_top_stance_position_combos(obs_with_kong, obs_without_kong, request=request)
    # -----------------------------------

    # ---Dog Breeds Distribution Chart---
    # Getting breeds and their counts
    dog_breeds = Dog.objects.filter(branch=branch).values('breed').annotate(total=Count('breed')).exclude(breed='').exclude(
        breed__isnull=True).order_by('-total')
    # -----------------------------------

    # ---Dog Stances by Day (Across The Week) Chart---
    top_dog_stances = get_top_dog_stances(request)
    top_dog_stances_limit = get_max_dog_stances_count(request)
    # Switch the Stances names to the front-end friendly version
    DOG_STANCE_DICT = dict(DogStance.DOG_STANCE_CHOICES)
    top_dog_stances_names = []
    for stance in top_dog_stances:
        top_dog_stances_names.append(DOG_STANCE_DICT[stance])

    stance_count_by_day = get_stance_count_by_day(top_dog_stances, request=request)

    # Set Up the yearly data for the yearly drop-down selection
    # Fetch the top stances per year and their limits
    top_stances_per_year, top_stances_per_year_limits = get_top_dog_stances_per_year(request)

    # Fetch the top stances per year dictionary for the chart dataset
    yearly_stance_count_by_day = get_yearly_stance_count_by_day(request)
    # -----------------------------------

    # Preparing data to be used in the frontend
    data = {
        'dogs_with_kong': serializers.serialize('json', dogs_with_kong),
        'stances_with_kong': list(stances_with_kong),  # Dog Stances With Kong (1) Chart
        'yearly_stances_with_kong': yearly_stances_with_kong,  # Dog Stances With Kong (2) Chart
        'years_with_kong': years_with_kong,  # Dog Stances With Kong (3) Chart (for drop-down yearly selection)
        'stances_without_kong': list(stances_without_kong),  # Dog Stances Without Kong (1) Chart
        'yearly_stances_without_kong': yearly_stances_without_kong,  # Dog Stances Without Kong (2) Chart
        'years_without_kong': years_without_kong,  # Dog Stances Without Kong (3) Chart (for drop-down yearly selection)
        'dog_breeds': list(dog_breeds),  # Dog Breeds Distribution Chart
        'stance_count_by_day': stance_count_by_day,  # Dog Stances by Day (Across The Week) Chart (1)
        'top_dog_stances': top_dog_stances_names,  # Dog Stances by Day (Across The Week) Chart (2)
        'top_dog_stances_limit': top_dog_stances_limit,  # Dog Stances by Day (Across The Week) Chart (3)
        'yearly_stance_count_by_day': yearly_stance_count_by_day,  # Dog Stances by Day (Across The Week) Chart (4)
        'top_stances_per_year': top_stances_per_year,  # Dog Stances by Day (Across The Week) Chart (5)
        'top_stances_per_year_limits': top_stances_per_year_limits,  # Dog Stances by Day (Across The Week) Chart (6)
        'top_stance_position_combos': top_stance_position_combos,  # Most Common General Behaviors Chart
        # Comprehensive Health & Safety Profile of Our Canines Chart (only for vets)
        'health_metrics': health_metrics if request.user.is_authenticated and (request.user.groups.filter(name='Vet').exists() or request.user.is_superuser) else None,
    }

    return JsonResponse(data)


# Switch the Stances names to the front-end friendly version
def map_stances_to_frontend(stances):
    # Prepare a dictionary for the front-end names
    DOG_STANCE_DICT = dict(DogStance.DOG_STANCE_CHOICES)

    for stance in stances:
        stance['dogStance'] = DOG_STANCE_DICT.get(stance['dogStance'], stance['dogStance'])
    return stances


# Fetch two dictionaries with the total counts of dogStances by years - with and without kong toys
# (for the ("Dog Stances With Kong" Chart)
def get_stances_count_by_year(request):
    branch = get_current_branch(request)

    # Get all Observations with and without a Kong Toy
    obs_with_kong = Observation.objects.filter(observes__dog__branch=branch, isKong='Y', obsDateTime__isnull=False)
    obs_without_kong = Observation.objects.filter(observes__dog__branch=branch, isKong='N', obsDateTime__isnull=False)

    # Get all DogStances associated with these Observations
    stances_with_kong = DogStance.objects.filter(observation__in=obs_with_kong)
    stances_without_kong = DogStance.objects.filter(observation__in=obs_without_kong)

    # Extract year in Python and count DogStances by year and type for the "Dog Stances With Kong" Chart
    stances_with_kong_grouped_by_year = {}
    for stance in stances_with_kong:
        year = stance.observation.obsDateTime.year if stance.observation.obsDateTime else None
        if year not in stances_with_kong_grouped_by_year:
            stances_with_kong_grouped_by_year[year] = {}

        dog_stance = stance.dogStance
        if dog_stance not in stances_with_kong_grouped_by_year[year]:
            stances_with_kong_grouped_by_year[year][dog_stance] = 0

        stances_with_kong_grouped_by_year[year][dog_stance] += 1

    # Extract year in Python and count DogStances by year and type for the "Dog Stances Without Kong" Chart
    stances_without_kong_grouped_by_year = {}
    for stance in stances_without_kong:
        year = stance.observation.obsDateTime.year if stance.observation.obsDateTime else None
        if year not in stances_without_kong_grouped_by_year:
            stances_without_kong_grouped_by_year[year] = {}

        dog_stance = stance.dogStance
        if dog_stance not in stances_without_kong_grouped_by_year[year]:
            stances_without_kong_grouped_by_year[year][dog_stance] = 0

        stances_without_kong_grouped_by_year[year][dog_stance] += 1

    # Convert the counts to the frontend format
    formatted_stances_with_kong = {
        year: map_stances_to_frontend([
            {'dogStance': stance, 'total': total}
            for stance, total in year_stances.items()
        ])
        for year, year_stances in stances_with_kong_grouped_by_year.items()
    }

    formatted_stances_without_kong = {
        year: map_stances_to_frontend([
            {'dogStance': stance, 'total': total}
            for stance, total in year_stances.items()
        ])
        for year, year_stances in stances_without_kong_grouped_by_year.items()
    }

    return formatted_stances_with_kong, formatted_stances_without_kong


# Returns the overall top dog stances
def get_top_dog_stances(request=None):
    branch = get_current_branch(request)

    # Get all the dogs in the current branch
    dogs = Dog.objects.filter(branch=branch)

    # Fetch and count the occurrences of each dogStance from the database for Dog Stances Across The Week
    dog_stance_counts = DogStance.objects.values('dogStance').filter(observation__observes__dog__in=dogs).annotate(total_count=Count('dogStance')).order_by(
        '-total_count')

    # Create a list to store the top X dogStances
    top_dog_stances = []

    # Collect the top X dogStances
    for i, entry in enumerate(dog_stance_counts):
        top_dog_stances.append(entry['dogStance'])

    return top_dog_stances


# Returns the top dog stances for each year and another dictionary for the limits of stances per year
def get_top_dog_stances_per_year(request):
    branch = get_current_branch(request)
    dogs = Dog.objects.filter(branch=branch)

    # Fetch all DogStance instances with their observation dates
    stances_with_datetime = DogStance.objects.filter(observation__observes__dog__in=dogs).select_related(
        'observation').values('observation__obsDateTime', 'dogStance')

    # Initialize a dictionary to group stances by year
    stances_by_year = {}

    # Loop through the stances and group them by year
    for entry in stances_with_datetime:
        # Extract the year from the observation datetime
        year = timezone.localtime(entry['observation__obsDateTime']).year

        # Append the stance to the corresponding year
        if year not in stances_by_year:
            stances_by_year[year] = []
        stances_by_year[year].append(entry['dogStance'])

    # Calculate top dog stances for each year
    top_stances_per_year = {}
    for year, stances in stances_by_year.items():
        stance_count = Counter(stances)
        top_stances = [stance for stance, _ in stance_count.most_common()]
        top_stances_per_year[year] = top_stances

    # Get the mapping from the DOG_STANCE_CHOICES
    stance_name_mapping = {backend: frontend for backend, frontend in DogStance.DOG_STANCE_CHOICES}

    # Replace back-end names with front-end friendly names
    for year, stances in top_stances_per_year.items():
        front_end_stances = [stance_name_mapping.get(stance, "Unknown") for stance in stances]
        top_stances_per_year[year] = front_end_stances

    # Initialize the dictionary to hold the stance limits for the slider per year
    top_stances_per_year_limits = {}

    # Loop through the top_dog_stances_per_year dictionary
    for year, stances in top_stances_per_year.items():
        # Calculate the length of the stances list and assign it to the corresponding year
        top_stances_per_year_limits[year] = len(stances)

    # Order the dictionary keys in descending order
    top_stances_per_year = dict(sorted(top_stances_per_year.items(), reverse=True))
    top_stances_per_year_limits = dict(sorted(top_stances_per_year_limits.items(), reverse=True))

    return top_stances_per_year, top_stances_per_year_limits


# Get a dictionary of each DogStance's occurrences in every day of the week, for a selected list of Stances
def get_stance_count_by_day(top_dog_stances, request):
    branch = get_current_branch(request)

    # Get all the dogs in the current branch
    dogs = Dog.objects.filter(branch=branch)

    days_of_week = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

    # Initialize a dictionary to hold the stance count for each day of the week
    stance_count_by_day = {}
    for day in days_of_week:
        stance_count_by_day[day] = {}
        for stance in top_dog_stances:
            stance_count_by_day[day][stance] = 0

    # Query the database to get the stances and their observation dates
    stances_with_datetime = DogStance.objects.filter(observation__observes__dog__in=dogs).select_related('observation').values('observation__obsDateTime', 'dogStance')
    # Mapping of Python's datetime.weekday() indexes to actual day names
    days_of_week_mapping = {0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 4: 'Friday', 5: 'Saturday',
                            6: 'Sunday'}

    # Mapping of value names in Dog Stances
    stance_name_mapping = {backend: frontend for backend, frontend in DogStance.DOG_STANCE_CHOICES}

    # Loop to populate the dictionary with actual data
    for entry in stances_with_datetime:
        # Convert the UTC time in the database to local time
        utc_observation_time = entry['observation__obsDateTime']
        local_observation_time = timezone.localtime(utc_observation_time)

        # Get the local weekday index (Monday is 0, Sunday is 6)
        day_index = local_observation_time.weekday()

        # Map the weekday index to its string name
        day_name = days_of_week_mapping[day_index]

        # Fetch the stance for this particular record
        stance = entry['dogStance']

        # If this stance is among the top stances we are tracking, update the count
        if stance in top_dog_stances:
            stance_count_by_day[day_name][stance] += 1

    transformed_stance_count_by_day = {}

    for day, stances in stance_count_by_day.items():
        transformed_stances = {}
        for stance_key, count in stances.items():
            transformed_key = stance_name_mapping.get(stance_key, "Unknown")
            transformed_stances[transformed_key] = count
        transformed_stance_count_by_day[day] = transformed_stances
    return transformed_stance_count_by_day


# Get a yearly dictionary of each DogStance's occurrences in every day of the week, for a selected list of Stances
def get_yearly_stance_count_by_day(request):
    branch = get_current_branch(request)
    dogs = Dog.objects.filter(branch=branch)

    # Fetch all DogStance instances with their observation dates
    stances_with_datetime = DogStance.objects.filter(observation__observes__dog__in=dogs).select_related('observation').values('observation__obsDateTime', 'dogStance')

    # Initialize a nested dictionary to hold the data
    stance_count_by_year_and_day = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    # Mapping of Python's datetime.weekday() indexes to actual day names
    days_of_week_mapping = {0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 4: 'Friday', 5: 'Saturday',
                            6: 'Sunday'}

    # Loop through the stances and group them by year and day
    for entry in stances_with_datetime:
        # Extract the year and day from the observation datetime
        observation_datetime = timezone.localtime(entry['observation__obsDateTime'])
        year = observation_datetime.year
        day_index = observation_datetime.weekday()
        day_name = days_of_week_mapping[day_index]

        # Fetch the stance for this particular record
        stance = entry['dogStance']

        # Increment the count for this stance in the specific year and day
        stance_count_by_year_and_day[year][day_name][stance] += 1

    # Convert defaultdicts to regular dicts
    stance_count_by_year_and_day = {year: {day: dict(stances) for day, stances in days.items()} for year, days in stance_count_by_year_and_day.items()}

    # Order the dictionary keys in descending order
    stance_count_by_year_and_day = dict(sorted(stance_count_by_year_and_day.items(), reverse=True))

    # Get the mapping from the DOG_STANCE_CHOICES
    stance_name_mapping = {backend: frontend for backend, frontend in DogStance.DOG_STANCE_CHOICES}

    # Replace back-end names with front-end friendly names
    for year, days in stance_count_by_year_and_day.items():
        for day, stances in days.items():
            front_end_stances = {stance_name_mapping.get(stance, "Unknown"): count for stance, count in stances.items()}
            stance_count_by_year_and_day[year][day] = front_end_stances

    # Convert defaultdicts to regular dicts
    stance_count_by_year_and_day = {year: {day: dict(stances) for day, stances in days.items()} for year, days in
                                    stance_count_by_year_and_day.items()}

    return stance_count_by_year_and_day


# Get the maximum number of stance options used in the DogStance in the current branch
# For the slider max limit above the "Top Dog Stances Across The Week" chart
def get_max_dog_stances_count(request):
    branch = get_current_branch(request)

    # Fetch all Dogs in the current branch
    dogs = Dog.objects.filter(branch=branch)

    # Fetch all DogStance objects in current branch
    all_stances = DogStance.objects.filter(observation__observes__dog__in=dogs)

    # Extract distinct dogStance values
    used_stances = set([stance.dogStance for stance in all_stances])

    return len(used_stances)


# Helper function for fetching a top 10 list of combined dogStances + dogPositions and their counts with/without kongs
def fetch_top_stance_position_combos(obs_with_kong, obs_without_kong, request):
    branch = get_current_branch(request)
    branch_dogs = Dog.objects.filter(branch=branch)

    # Define the mapping dictionaries for the final display
    stance_choices_dict = dict(DogStance.DOG_STANCE_CHOICES)
    location_choices_dict = dict(DogStance.DOG_LOCATION_CHOICES)

    # Fetch a list of top stance+location combinations with kong
    stance_pos_combo_with = DogStance.objects.filter(observation__observes__dog__in=branch_dogs,
                                                     observation__in=obs_with_kong).annotate(stance_location=Concat('dogStance',
                                                                                                                    Value(' + '),
                                                                                                                    'dogLocation',
                                                                                                                    output_field=CharField())).values('stance_location').annotate(count=Count('stance_location')).order_by('-count')[:10]

    # Fetch a list of top stance+location combinations with kong
    stance_pos_combo_without = DogStance.objects.filter(observation__observes__dog__in=branch_dogs,
                                                        observation__in=obs_without_kong).annotate(stance_location=Concat('dogStance',
                                                                                                                          Value(' + '),
                                                                                                                          'dogLocation',
                                                                                                                          output_field=CharField())).values('stance_location').annotate(count=Count('stance_location')).order_by('-count')[:10]

    # Replace the values in the QuerySets for both lists
    for stance in stance_pos_combo_with:
        stance_db, location_db = stance['stance_location'].split(' + ')
        location_display = location_choices_dict.get(location_db, '')
        stance[
            'stance_location'] = f"{stance_choices_dict[stance_db]} {location_display}".strip()

    for stance in stance_pos_combo_without:
        stance_db, location_db = stance['stance_location'].split(' + ')
        location_display = location_choices_dict.get(location_db, '')
        stance[
            'stance_location'] = f"{stance_choices_dict[stance_db]} {location_display}".strip()

    # Union of both lists keys
    unique_keys = set([item['stance_location'] for item in stance_pos_combo_with] + [item['stance_location'] for item in
                                                                                     stance_pos_combo_without])

    # Initialize dictionaries with zeros
    dict_with = {key: 0 for key in unique_keys}
    dict_without = {key: 0 for key in unique_keys}

    # Populate the dictionaries with actual counts
    for item in stance_pos_combo_with:
        dict_with[item['stance_location']] = item['count']

    for item in stance_pos_combo_without:
        dict_without[item['stance_location']] = item['count']

    # Create a list containing the unique combinations and their counts in both lists
    combined_list = [(key, dict_with[key], dict_without[key]) for key in unique_keys]

    # Sort the list based the sum of both counts for sorting
    combined_list.sort(key=lambda x: x[1] + x[2], reverse=True)

    # Return the combined top 10 Stance+Position combos and the counts for both with and without kong
    return combined_list[:10]


# Helper function to filter out unwanted attributes.
def exclude_unwanted(attribute):
    exclusions = ["Unspecified", " ", "-", ""]
    return Q(**{f"{attribute}__isnull": True}) | Q(**{f"{attribute}__in": exclusions})


# Fetch unique values for a given dog attribute.
def get_unique_values(request, attribute):
    branch = get_current_branch(request)

    return (Dog.objects.filter(owner__branch=branch)
            .exclude(exclude_unwanted(attribute))
            .values(attribute).distinct().order_by(attribute))


# Fetch unique owner values.
def get_unique_owners(request):
    branch = get_current_branch(request)

    return (Dog.objects.filter(owner__branch=branch)
            .exclude(owner__isnull=True)
            .values('owner__firstName',
                    'owner__lastName',
                    'owner').distinct().order_by('owner__firstName',
                                                 'owner__lastName'))


# Prepare a dictionary in JSON for distribution of dogs
# by gender,  vaccination and isneutered. Used for health_metrics_chart.
def get_health_metrics_dict(request):
    branch = get_current_branch(request)
    # Fetch required attributes of all dogs
    dogs_data = Dog.objects.filter(branch=branch).values_list('gender', 'dateOfVaccination', 'isNeutered', 'isDangerous')

    # Initialize data structure
    health_metrics_dict = {
        'gender': {'M': 0, 'F': 0},
        'vaccinated': {'M': {'Y': 0, 'N': 0}, 'F': {'Y': 0, 'N': 0}},
        'neutered': {
            'M': {'Y': {'Y': 0, 'N': 0, '-': 0}, 'N': {'Y': 0, 'N': 0, '-': 0}},
            'F': {'Y': {'Y': 0, 'N': 0, '-': 0}, 'N': {'Y': 0, 'N': 0, '-': 0}}
        }
    }

    # Loop through each dog entry to populate chart_data
    for gender, dateOfVaccination, isNeutered, isDangerous in dogs_data:
        # Skip dogs with unknown gender
        if gender not in ['M', 'F']:
            continue

        # Calculate if vaccination is within the last 365 days or not, if null set as No as well
        if dateOfVaccination:
            current_date = timezone.now().date()
            delta = current_date - dateOfVaccination
            isVaccinated = 'Y' if delta.days <= 365 else 'N'
        else:
            isVaccinated = 'N'

        # Handle Nulls and empties
        isNeutered = isNeutered or '-'

        # Populate the counters
        health_metrics_dict['gender'][gender] += 1
        health_metrics_dict['vaccinated'][gender][isVaccinated] += 1
        health_metrics_dict['neutered'][gender][isVaccinated][isNeutered] += 1

    return health_metrics_dict


# View for displaying all News in a dedicated news page
def view_news(request):
    # Get the current Branch
    branch = get_current_branch(request)

    news_list = News.objects.filter(branch=branch).order_by('-created_at')  # Fetch news in descending order
    context = {'news_list': news_list}
    return render(request, 'news_page.html', context)


@regular_user_required
# View for viewing all dogs in a table
def view_dogs(request):
    branch = get_current_branch(request)

    # Check if user is logged in
    if request.user.is_authenticated:
        # Apply sorting by attributes
        sort_by = request.GET.get('sort_by', '-dateOfArrival')

        # Fetch all filtered dogs
        dog_filter = DogFilter(request.GET, queryset=Dog.objects.filter(branch=branch).order_by(sort_by))
        filtered_dogs = dog_filter.qs

        # Prepare a list of unique breeds, fur colors and owners, exclude redundant results
        unique_breeds = get_unique_values(request, 'breed')
        unique_colors = get_unique_values(request, 'furColor')
        unique_owners = get_unique_owners(request)

        # Pagination logic
        paginator = Paginator(filtered_dogs, ENTRIES_PER_PAGE)
        page = request.GET.get('page', 1)
        dogs_page = paginator.get_page(page)
        pagination_start = max(dogs_page.number - 3, 1)
        pagination_end = min(dogs_page.number + 3, paginator.num_pages)

        context = {
            'dogs': dogs_page,
            'unique_breeds': unique_breeds,
            'unique_colors': unique_colors,
            'unique_owners': unique_owners,
            'pagination_start': pagination_start,
            'pagination_end': pagination_end,
        }

        return render(request, 'view_dogs.html', context=context)
    # If user is not logged in/authenticated, show an error message and redirect to home.
    else:
        messages.error(request, "You Must Be Logged In To View Dogs...")
        return redirect('dogs_app:home')


# Helper function to filter dogs based on date range.
def date_filter_logic(dogs, start_date, end_date, field_name):
    if start_date and end_date:
        dogs = dogs.filter(**{f'{field_name}__range': [start_date, end_date]})
    elif start_date:
        dogs = dogs.filter(**{f'{field_name}__gte': start_date})
    elif end_date:
        dogs = dogs.filter(**{f'{field_name}__lte': end_date})
    return dogs


# Helper function to filter dogs based on 'Unspecified' value.
def unspecified_filter(dogs, field_name, value):
    if value == "Unspecified":
        dogs = dogs.filter(Q(**{f'{field_name}__exact': ''}) | Q(**{f'{field_name}__isnull': True}))
    return dogs


@regular_user_required
# Handling Dog Filtering in the view_dogs page
def filter_dogs(request):
    branch = get_current_branch(request)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Apply sorting by attributes
        sort_by = request.GET.get('sort_by', '-dateOfArrival')

        # Initialize queryset
        dogs = Dog.objects.filter(branch=branch).order_by(sort_by)

        # Date filtering
        for field in ['dateOfArrival', 'dateOfVaccination', 'kongDateAdded']:
            # Handle "kongDateAdded" having a capital D
            adjusted_field = field if "kong" not in field else field.replace('Date', 'date', 1)

            start = request.GET.get(adjusted_field.replace('date', 'startDate'), None)
            end = request.GET.get(adjusted_field.replace('date', 'endDate'), None)
            dogs = date_filter_logic(dogs, start, end, field)

        # Retrieve dog age range from the request, handle empty cases
        age_from_str = request.GET.get('ageFrom', '0')
        min_age = float(age_from_str) if age_from_str else 0.0

        age_to_str = request.GET.get('ageTo', '20')
        max_age = float(age_to_str) if age_to_str else 20.0

        # Convert age to date range
        end_date_birth = date.today() - timedelta(days=min_age * 365)
        start_date_birth = date.today() - timedelta(days=max_age * 365)

        # If the min and max age are default values, include dogs without age attribute as well.
        if min_age == 0 and max_age == 20:
            dogs = dogs.filter(
                Q(dateOfBirthEst__range=[start_date_birth, end_date_birth]) | Q(dateOfBirthEst__isnull=True))
        else:
            # Filter by age range
            dogs = dogs.filter(dateOfBirthEst__range=[start_date_birth, end_date_birth])

        # Apply Django Filters
        filtered_dogs = DogFilter(request.GET, queryset=dogs)
        dogs = filtered_dogs.qs

        # Check for "Unspecified" conditions and apply those filters.
        unspecified_fields = ['breed', 'gender', 'furColor', 'isNeutered', 'isDangerous']
        for field in unspecified_fields:
            value = request.GET.get(field, None)
            dogs = unspecified_filter(dogs, field, value)

        # Check for "Unspecified" in Owner separately, apply filters if needed.
        owner = request.GET.get('owner', None)
        if owner == "Unspecified":
            dogs = dogs.filter(Q(owner__isnull=True))

        # Store all dog IDs being displayed in case user wants to export the data using get_filtered_dogs_id()
        filtered_dogs_ids = list(dogs.values_list('dogID', flat=True))
        request.session['filtered_dogs_ids'] = filtered_dogs_ids

        # Pagination
        paginator = Paginator(dogs, ENTRIES_PER_PAGE)
        page = request.GET.get('page')
        dogs_page = paginator.get_page(page)
        pagination_start = max(dogs_page.number - 3, 1)
        pagination_end = min(dogs_page.number + 3, paginator.num_pages)

        # Render table and pagination
        table_rows = ''.join([render_to_string('_dog_row.html', {'dog': dog, 'user': request.user}) for dog in dogs_page])
        pagination_html = render_to_string('_pagination.html', {
            'dogs': dogs_page,
            'pagination_start': pagination_start,
            'pagination_end': pagination_end,
        })

        return JsonResponse({'table_rows': table_rows, 'pagination_html': pagination_html}, status=200)
    else:
        return JsonResponse({'error': 'Not an AJAX request'}, status=400)


# Helper function for getting all dog IDs for dogs being filtered on the page (for exporting data in JSON/Excel)
def get_filtered_dog_ids(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        filtered_dogs_ids = request.session.get('filtered_dogs_ids', [])
        return JsonResponse({'filtered_dogs_ids': filtered_dogs_ids}, status=200)
    else:
        return JsonResponse({'error': 'Not an AJAX request'}, status=400)


# Main function to export Dogs in JSON format
def export_dogs_json(request):
    if request.user.is_authenticated and (request.user.is_superuser or request.user.groups.filter(name='Vet').exists()):
        if request.method == 'POST':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                dog_ids = json.loads(request.POST.get('dog_ids'))

                # Use select_related and prefetch_related for optimization and fetching all associated entities
                dogs = Dog.objects \
                    .select_related('owner') \
                    .prefetch_related('entranceexamination_set',
                                      'treatment_set',
                                      'dogplacement_set__kennel',
                                      Prefetch('observers',
                                               queryset=Observes.objects.prefetch_related(
                                                   Prefetch('observation_set',
                                                            queryset=Observation.objects
                                                            .prefetch_related('dogstance_set'))))) \
                    .filter(dogID__in=dog_ids)

                # Serialize dogs
                dog_data = DogSerializer.serialize_dogs(dogs)

                # Add a root key "data" to wrap the list
                wrapped_dog_data = {'data': dog_data}

                return JsonResponse(wrapped_dog_data, safe=False)
    else:
        # If user is not logged in/authenticated, show an error message and redirect to home.
        return JsonResponse({'error': 'You are not authorized to perform this action'}, status=403)


# View for handling Excel file exports in view_dogs page
def export_dogs_excel(request):
    if request.user.is_authenticated and (request.user.is_superuser or request.user.groups.filter(name='Vet').exists()):
        if request.method == 'POST':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                try:
                    data = json.loads(request.body.decode('utf-8'))
                    dog_ids = data.get('dog_ids', [])
                    sort_by = data.get('sort_by', '-dateOfArrival')

                    # Filter and sort dogs
                    dogs = Dog.objects.filter(dogID__in=dog_ids).order_by(sort_by)

                    # Create workbook and worksheet
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Dogsheler Dogs Data"

                    # Define headers and write them to the first row
                    headers = ['Chip Number', 'Name', 'Date of Birth', 'Date of Arrival',
                               'Date of Vaccination', 'Breed', 'Gender', 'Fur Color', 'Neutered',
                               'Dangerous', 'Image', 'Last Kong Date Given', 'Owner']

                    # Styling headers with bold font and background color
                    header_font = Font(bold=True, color="FFFFFF")
                    data_font = Font(name='Arial', size=11)

                    # Define fills
                    header_fill = PatternFill(start_color="0070C0",
                                              end_color="0070C0", fill_type="solid")
                    light_blue_fill = PatternFill(start_color="D9EBF5", end_color="D9EBF5", fill_type="solid")
                    light_gray_fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")

                    # Define Center alignment
                    center_aligned = Alignment(horizontal="center", vertical="center")

                    # Define border
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

                    # Header row styling
                    for col_num, header in enumerate(headers, 1):
                        col_letter = get_column_letter(col_num)
                        cell = ws['{}1'.format(col_letter)]
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_aligned
                        cell.border = thin_border
                        ws.column_dimensions[col_letter].width = 15 if header not in ['Gender', 'Neutered',
                                                                                      'Dangerous'] else 10

                    # Check if dogs queryset is empty
                    if not dogs.exists():
                        # Merge cells from A2 to N2 (14 columns)
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)

                        # Reference to the merged cell
                        merged_cell = ws.cell(row=2, column=1)

                        # Populate the merged cell
                        merged_cell.value = "No Data Available"

                        # Style the merged cell
                        merged_cell.font = Font(name='Arial', size=12, bold=True)
                        merged_cell.alignment = center_aligned
                        merged_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                        merged_cell.border = thin_border

                    # Otherwise, proceed with populating Excel file with dog data and apply styles
                    else:
                        # Populate Excel file with dog data and apply styles
                        for row_num, dog in enumerate(dogs, 2):  # Start from row 2 to not overwrite headers
                            for col_num in range(1, 14):  # 13 columns in total
                                # Determine row color
                                if row_num % 2 == 0:
                                    row_fill = light_gray_fill
                                else:
                                    row_fill = light_blue_fill

                                cell = ws.cell(row=row_num, column=col_num)

                                # Populate the cell based on the column number
                                cell.value = {
                                    1: dog.chipNum if dog.chipNum else "N/A",
                                    2: dog.dogName if dog.dogName else "N/A",
                                    3: dog.dateOfBirthEst if dog.dateOfBirthEst else "N/A",
                                    4: dog.dateOfArrival if dog.dateOfArrival else "N/A",
                                    5: dog.dateOfVaccination if dog.dateOfVaccination else "N/A",
                                    6: dog.breed if dog.breed else "N/A",
                                    7: dog.get_gender_display() if dog.gender else "N/A",
                                    8: dog.furColor if dog.furColor else "N/A",
                                    9: dog.get_isNeutered_display() if dog.isNeutered else "N/A",
                                    10: dog.get_isDangerous_display() if dog.isDangerous else "N/A",
                                    11: str(dog.dogImage) if dog.dogImage else "N/A",
                                    12: dog.kongDateAdded if dog.kongDateAdded else "N/A",
                                    13: str(dog.owner) if dog.owner else "N/A"
                                }.get(col_num)

                                # Apply styling
                                cell.font = data_font
                                cell.alignment = center_aligned
                                cell.border = thin_border
                                cell.fill = row_fill

                    # Initialize BytesIO and save workbook to it
                    excel_file = BytesIO()
                    wb.save(excel_file)
                    excel_file.seek(0)

                    # Prepare the HttpResponse
                    response = HttpResponse(
                        excel_file.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=dogs_data.xlsx'

                    return response
                except Exception as e:
                    # Handling any exception that occurs during export
                    error_message = f"Export failed due to error: {str(e)}"
                    return JsonResponse({'status': 'error', 'message': error_message}, status=400)
            else:
                return JsonResponse({'error': 'Not an AJAX request'}, status=400)
        else:
            return JsonResponse({'error': 'Invalid request method'}, status=400)
    else:
        # If user is not logged in/authenticated, show an error message and redirect to home.
        return JsonResponse({'error': 'You are not authorized to perform this action'}, status=403)


# View for handling Excel file imports in view_dogs page
def import_dogs_excel(request):
    if request.user.is_authenticated and (request.user.is_superuser or request.user.groups.filter(name='Vet').exists()):
        if request.method == 'POST':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                try:
                    excel_file = request.FILES['excel_file']
                    wb = load_workbook(excel_file)
                    ws = wb.active

                    dog_count = 0  # Initialize a counter for imported dogs

                    # Start the transaction block
                    with transaction.atomic():
                        # Iterate through each row in the worksheet
                        for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
                            try:
                                # Validate Data First
                                gender_value = 'M'
                                is_neutered_value = ''
                                is_dangerous_value = ''

                                if row[6].value:
                                    gender_value = 'M' if row[6].value.lower() in ['male', 'm'] or row[6].value in ['זכר',
                                                                                                                    'ז'] \
                                        else 'F' if row[6].value.lower() in ['female', 'f'] or row[6].value in ['נקבה', 'נ'] \
                                        else 'M'
                                if row[8].value:
                                    is_neutered_value = 'Y' if row[8].value.lower() in ['yes', 'y'] or row[8].value in [
                                        'כן', 'כ'] \
                                        else 'N' if row[8].value.lower() in ['no', 'n'] or row[8].value in ['לא', 'ל'] \
                                        else ''
                                if row[9].value:
                                    is_dangerous_value = 'Y' if row[9].value.lower() in ['yes', 'y'] or row[9].value in [
                                        'כן', 'כ'] \
                                        else 'N' if row[9].value.lower() in ['no', 'n'] or row[9].value in ['לא', 'ל'] \
                                        else ''

                                date_of_arrival_value = row[3].value
                                if not date_of_arrival_value:
                                    date_of_arrival_value = timezone.now().date()

                                # dog_image_url = None if row[10].value is None else '/static/img/' + row[10].value
                                dog_image_url = None  # TO-DO Work on Images

                                dog_data = {
                                    'chipNum': row[0].value,
                                    'dogName': row[1].value,
                                    'dateOfBirthEst': row[2].value,
                                    'dateOfArrival': date_of_arrival_value,
                                    'dateOfVaccination': row[4].value,
                                    'breed': row[5].value,
                                    'gender': gender_value,
                                    'furColor': row[7].value,
                                    'isNeutered': is_neutered_value,
                                    'isDangerous': is_dangerous_value,
                                    'dogImage': dog_image_url,
                                    'kongDateAdded': row[11].value,
                                    'owner': None
                                    # 'owner': omitted for now
                                }

                                # Attempt to create a new Dog instance
                                new_dog = Dog(**dog_data)
                                new_dog.full_clean()  # This will raise ValidationError for any field issues
                                new_dog.save()
                                # Increment dog count
                                dog_count += 1

                            except ValidationError as e:
                                # Handling specific field validation errors
                                error_details = "\n".join(
                                    [f"Column '{k}': {', '.join(v)}" for k, v in e.message_dict.items()])
                                error_message = f"Import error at row {row_number}:{error_details}"
                                raise ValueError(error_message)

                            except IntegrityError as e:
                                # Handling unique constraint violations like chipNum uniqueness
                                error_message = f"Import error at row {row_number}:{str(e)}"
                                raise ValueError(error_message)

                        # add success message with dog count
                        if dog_count == 1:
                            success_message = "1 dog has been successfully imported!"
                        else:
                            success_message = f"{dog_count} dogs have been successfully imported!"
                        return JsonResponse({'status': 'success', 'message': success_message}, status=200)
                except ValueError as e:
                    # Handling custom ValueError which now includes detailed row information
                    error_message = f"Import cancelled due to error:{e}"
                    return JsonResponse({'status': 'error', 'message': error_message}, status=400)

                except Exception as e:
                    # General error handling
                    general_error_message = f"General error occurred: {str(e)}. Import cancelled."
                    return JsonResponse({'status': 'error', 'message': general_error_message}, status=400)
            else:
                # If request is not AJAX, return error
                return JsonResponse({'error': 'Not an AJAX request'}, status=400)
        else:
            # If request method is not POST, return error
            return JsonResponse({'error': 'Invalid request method'}, status=400)
    else:
        return JsonResponse({'error': 'You are not authorized to perform this action'}, status=403)


# View for handling JSON file imports in view_dogs page
def import_dogs_json(request):
    if request.user.is_authenticated and (request.user.is_superuser or request.user.groups.filter(name='Vet').exists()):
        branch = get_current_branch(request)
        if request.method == 'POST':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                try:
                    json_file = request.FILES['json_file']
                    data = json.load(json_file)['data']

                    dog_count = 1  # Initialize a counter for imported dogs

                    # Start the transaction block
                    with transaction.atomic():
                        # Iterate through each dog in the data
                        for index, dog_info in enumerate(data):
                            try:
                                # Remove dogID from the data, hold the rest of the associated entities
                                dog_info.pop('dogID', None)

                                # Hold the rest of the associated entities
                                owner_info = dog_info.pop('owner', None)
                                treatments_info = dog_info.pop('treatments', [])
                                examinations_info = dog_info.pop('entranceExaminations', [])
                                dogPlacements_info = dog_info.pop('dogPlacements', [])
                                observes_info = dog_info.pop('observes', [])

                                # Default 'dateOfArrival' to today if not provided
                                if not dog_info.get('dateOfArrival'):
                                    dog_info['dateOfArrival'] = timezone.now().date().isoformat()

                                # Handle 'owner' field
                                owner = None  # In case we don't need to associate Dog with an Owner
                                if owner_info:
                                    owner_info.pop('ownerSerialNum', None)  # Remove ownerSerialNum which is PK auto field
                                    owner_id = owner_info.get('ownerID', None)

                                    if owner_id:
                                        # If ownerID is provided, use get_or_create
                                        owner, created = Owner.objects.get_or_create(
                                            ownerID=owner_id,
                                            defaults={k: v for k, v in owner_info.items()}
                                        )

                                        if not created:
                                            # If owner already exists, update the owner info
                                            for k, v in owner_info.items():
                                                setattr(owner, k, v)
                                            owner.branch = branch
                                            owner.save()
                                    else:
                                        # ownerID not provided, create an owner without one
                                        owner = Owner(**owner_info)
                                        owner.branch = branch
                                        owner.save()

                                # Extract and validate dog data from dog_info
                                # Create Dog instance and save
                                new_dog = Dog(**dog_info, owner=owner)
                                new_dog.full_clean()
                                new_dog.save()

                                # Handle 'treatment' field
                                treatment_count = 0  # Initialize counter for treatments
                                for treatment_info in treatments_info:
                                    treatment_count += 1
                                    try:
                                        treatment_info.pop('treatmentID', None)
                                        new_treatment = Treatment(**treatment_info, dog=new_dog)
                                        new_treatment.full_clean()  # Validate the treatment data
                                        new_treatment.save()
                                    except ValidationError as e:
                                        error_details = "; ".join([f"{field}: {', '.join(err_msgs)}" for field, err_msgs in
                                                                   e.message_dict.items()])
                                        error_message = (f"Dog #{index + 1}, Treatment #{treatment_count},"
                                                         f" Details: {error_details}")
                                        raise ValueError(error_message)

                                # Handle 'entranceExamination' field
                                examination_count = 0  # Initialize counter for examinations
                                for examination_info in examinations_info:
                                    examination_count += 1
                                    try:
                                        examination_info.pop('examinationID', None)
                                        new_examination = EntranceExamination(**examination_info, dog=new_dog)
                                        new_examination.full_clean()
                                        new_examination.save()
                                    except ValidationError as e:
                                        error_details = "; ".join([f"{field}: {', '.join(err_msgs)}" for field, err_msgs in
                                                                   e.message_dict.items()])
                                        error_message = (f"Dog #{index + 1}, Examination #{examination_count},"
                                                         f" Details: {error_details}")
                                        raise ValueError(error_message)

                                # Handle 'dogPlacement' field
                                dogPlacement_count = 0  # Initialize counter for placements
                                for dogPlacement_info in dogPlacements_info:
                                    dogPlacement_count += 1
                                    try:
                                        kennel_num = dogPlacement_info.pop('kennelNum', None)
                                        kennel_image = dogPlacement_info.pop('kennelImage', None)

                                        kennel = None

                                        if kennel_num:
                                            kennel, created = Kennel.objects.get_or_create(
                                                kennelNum=kennel_num,
                                                defaults={'kennelImage': kennel_image}
                                            )

                                            if not created:
                                                # If kennel already exists, update the kennel info
                                                kennel.kennelImage = kennel_image
                                                kennel.branch = branch
                                                kennel.save()

                                        new_dogPlacement = DogPlacement(**dogPlacement_info, dog=new_dog, kennel=kennel)
                                        new_dogPlacement.full_clean()
                                        new_dogPlacement.save()
                                    except ValidationError as e:
                                        error_details = "; ".join([f"{field}: {', '.join(err_msgs)}" for field, err_msgs in
                                                                   e.message_dict.items()])
                                        error_message = (f"Dog #{index + 1}, DogPlacement #{dogPlacement_count},"
                                                         f" Details: {error_details}")
                                        raise ValueError(error_message)

                                # Handle 'observes' field
                                observes_count = 0  # Initialize counter for Observes
                                for observes_data in observes_info:
                                    observes_count += 1
                                    try:
                                        camID = observes_data.pop('camID', None)
                                        observations_info = observes_data.pop('observations', [])

                                        # Default 'sessionDate' to today if not provided
                                        if not observes_data.get('sessionDate'):
                                            observes_data['sessionDate'] = timezone.now().date().isoformat()

                                        camera = None
                                        if camID:
                                            camera, created = Camera.objects.get_or_create(camID=camID, branch=branch)  # TO-DO: Check if this is correct

                                        new_observes = Observes(**observes_data, dog=new_dog, camera=camera)
                                        new_observes.full_clean()
                                        new_observes.save()
                                    except ValidationError as e:
                                        error_details = "; ".join([f"{field}: {', '.join(err_msgs)}" for field, err_msgs in
                                                                   e.message_dict.items()])
                                        error_message = (f"Dog #{index + 1}, Observes(Session) #{observes_count},"
                                                         f" Details: {error_details}")
                                        raise ValueError(error_message)

                                    # Handle 'observation' field
                                    observation_count = 0  # Initialize counter for Observation
                                    for observation_data in observations_info:
                                        observation_count += 1
                                        try:
                                            obs_date_str = observation_data.get('obsDateTime')

                                            # Check if obsDateTime is a string that needs parsing
                                            if obs_date_str:
                                                # Parse the string to a naive datetime object
                                                naive_datetime = parse_datetime(obs_date_str)
                                                if naive_datetime and not timezone.is_aware(naive_datetime):
                                                    # Localize the naive datetime
                                                    local_tz = pytz.timezone('Asia/Jerusalem')
                                                    aware_datetime = local_tz.localize(naive_datetime)
                                                    observation_data['obsDateTime'] = aware_datetime
                                            else:
                                                # Default 'obsDateTime' to now if not provided
                                                observation_data['obsDateTime'] = timezone.now()

                                            dogStances_info = observation_data.pop('dogStances', [])

                                            new_observation = Observation(**observation_data, observes=new_observes)
                                            new_observation.full_clean()
                                            new_observation.save()
                                        except ValidationError as e:
                                            error_details = "; ".join(
                                                [f"{field}: {', '.join(err_msgs)}" for field, err_msgs in
                                                 e.message_dict.items()])
                                            error_message = (f"Dog #{index + 1}, Observation #{observation_count},"
                                                             f" Details: {error_details}")
                                            raise ValueError(error_message)

                                        # Handle 'dogStance' field
                                        dogStance_count = 0  # Initialize counter for dogStance
                                        for dogStance_data in dogStances_info:
                                            dogStance_count += 1
                                            try:
                                                new_dogStance = DogStance(**dogStance_data, observation=new_observation)
                                                new_dogStance.full_clean()
                                                new_dogStance.save()
                                            except ValidationError as e:
                                                error_details = "; ".join(
                                                    [f"{field}: {', '.join(err_msgs)}" for field, err_msgs in
                                                     e.message_dict.items()])
                                                error_message = (f"Dog #{index + 1}, dogStance #{dogStance_count},"
                                                                 f" Details: {error_details}")
                                                raise ValueError(error_message)

                            except ValidationError as e:
                                # Handling specific field validation errors
                                error_details = "\n".join(
                                    [f"Field '{k}': {', '.join(v)}" for k, v in e.message_dict.items()])
                                error_message = f"Dog #{index + 1} (JSON Line: {index + 1}), {error_details}"
                                raise ValueError(error_message)

                            except IntegrityError as e:
                                # Handling unique constraint violations like chipNum uniqueness
                                error_message = f"Dog #{index + 1} (JSON Line: {index + 1}), {str(e)}"
                                raise ValueError(error_message)

                            dog_count += 1  # Increment after each successful processing

                        # add success message with dog count
                        if dog_count == 1:  # No dogs were successfully imported
                            success_message = ("Import Completed: No new dogs were added. "
                                               "Please check the data for accuracy or duplication.")
                        elif dog_count == 2:  # Since we start count from 1, 2 means only 1 dog processed
                            success_message = "1 dog has been successfully imported!"
                        else:
                            success_message = f"{dog_count - 1} dogs have been successfully imported!"
                        return JsonResponse({'status': 'success', 'message': success_message}, status=200)
                except ValueError as e:
                    # Handling custom ValueError which now includes detailed row information
                    error_message = f"Import Cancelled: Error Detected - {e}"
                    return JsonResponse({'status': 'error', 'message': error_message}, status=400)

                except Exception as e:
                    # General error handling
                    general_error_message = f"General error occurred: {str(e)}. Import cancelled."
                    return JsonResponse({'status': 'error', 'message': general_error_message}, status=400)
            else:
                # If request is not AJAX, return error
                return JsonResponse({'error': 'Not an AJAX request'}, status=400)
        else:
            # If request method is not POST, return error
            return JsonResponse({'error': 'Invalid request method'}, status=400)
    else:
        return JsonResponse({'error': 'You are not authorized to perform this action'}, status=403)


@login_required
def vote(request, poll_id):
    if 'voted_polls' in request.session:
        if str(poll_id) in request.session['voted_polls']:
            return HttpResponseBadRequest("You have already voted in this poll.")
    else:
        request.session['voted_polls'] = []

    if request.method == 'POST':
        poll_id = request.POST.get('poll_id')
        choice_id = request.POST.get('choice')
        if poll_id and choice_id:
            try:
                choice = Choice.objects.get(pk=choice_id)
            except Choice.DoesNotExist:
                return redirect('dogs_app:home')  # Redirect to home page if choice does not exist
            else:
                choice.votes += 1
                choice.save()
        # After successful voting, add the poll ID to the voted_polls session variable
        request.session['voted_polls'].append(str(poll_id))
        request.session.modified = True  # Ensure the session is saved

        # Set a cookie to remember that the user has voted in this session
        response = HttpResponse("Vote successful")
        expiration_date = timedelta(days=365)  # 1 year
        response.set_cookie('voted_poll_' + str(poll_id), 'true', max_age=expiration_date.total_seconds())
        messages.success(request, 'Thank you for voting!')

    return redirect('dogs_app:home')


def results(request, poll_id):
    poll = get_object_or_404(Poll, pk=poll_id)
    return render(request, 'polls/results.html', {'poll': poll})


@superuser_required
# View for adding a poll to the homepage. Only logged-in admins permitted
def add_poll(request):
    ChoiceFormSet = formset_factory(ChoiceForm, extra=4)  # Adjust extra as needed
    if request.method == 'POST':
        branch = get_current_branch(request)
        poll_form = PollForm(request.POST)
        choice_formset = ChoiceFormSet(request.POST)

        if poll_form.is_valid() and choice_formset.is_valid():
            poll = poll_form.save(commit=False)
            poll.branch = branch
            poll.save()
            
            for form in choice_formset:
                choice_text = form.cleaned_data.get('choice_text')
                if choice_text:
                    Choice.objects.create(poll=poll, choice_text=choice_text)

            messages.success(request, 'Poll has been added successfully!')
            return redirect('dogs_app:home')
    else:
        poll_form = PollForm()
        choice_formset = ChoiceFormSet()

    context = {
        'poll_form': poll_form,
        'choice_formset': choice_formset,
    }
    return render(request, 'add_poll.html', context)


@superuser_required
# View for updating a Poll from the homepage. Only available to Admins
def update_poll(request, poll_id):
    # Get the poll object or return a 404 error if it doesn't exist
    poll = get_object_or_404(Poll, pk=poll_id)

    # Get the current Branch
    branch = get_current_branch(request)  # Make sure to define get_current_branch function
    # Check if the poll belongs to the current branch
    if poll.branch != branch:
        messages.error(request, "You must be in the correct branch to delete this poll...")
        return redirect('dogs_app:home')

    if request.method == 'POST':
        # Create a form instance and populate it with data from the request (binding):
        poll_form = PollForm(request.POST, instance=poll)
        choice_formset = ChoiceFormSet(request.POST, instance=poll)

        # Check if both forms are valid
        if poll_form.is_valid() and choice_formset.is_valid():
            # Save the form data to the database
            poll_form.save()
            choice_formset.save()

            # Redirect to the homepage or any other page after updating the poll
            return redirect('dogs_app:home')  # Change 'home' to the appropriate URL name

    else:
        # Create a form instance and populate it with initial data for the poll
        poll_form = PollForm(instance=poll)
        choice_formset = ChoiceFormSet(instance=poll)

    # Render the update_poll.html template with the form instances
    return render(request, 'update_poll.html', {'poll_form': poll_form, 'choice_formset': choice_formset})


@superuser_required
# View for deleting a Poll from the homepage. Only available to Admins
def delete_poll(request, poll_id):
    # Get the poll to delete
    poll = get_object_or_404(Poll, pk=poll_id)
    # Get the current Branch
    branch = get_current_branch(request)  # Make sure to define get_current_branch function
    # Check if the poll belongs to the current branch
    if poll.branch != branch:
        messages.error(request, "You must be in the correct branch to delete this poll...")
        return redirect('dogs_app:home')
    if request.method == 'POST':
        poll.delete()
        messages.success(request, f"Poll: '{poll.question}' has been successfully deleted...")
        return redirect('dogs_app:home')
    else:
        return render(request, 'delete_poll.html', {'poll': poll})


@regular_user_required
# View for showing Polls
def view_polls(request):
    branch = get_current_branch(request)

    polls = Poll.objects.all().filter(branch=branch)
    return render(request, 'polls_page.html', {'polls': polls})

    

